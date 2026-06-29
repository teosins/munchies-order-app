import streamlit as st
import pandas as pd
import math
import io
import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Fillrate", page_icon="📦", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── global font ── */
html, body, [class*="css"], .stApp, button, input, select, textarea {
    font-family: 'Inter', sans-serif !important;
}

/* ── hide default streamlit chrome ── */
#MainMenu, footer, header {visibility: hidden;}

/* ── app background ── */
.stApp {
    background: #0a0a12;
}

/* ── main content padding ── */
.block-container {
    padding-top: 2rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
}

/* ── sidebar ── */
[data-testid="stSidebar"] {
    background: #0d0d1a !important;
    border-right: 1px solid rgba(255,255,255,0.06);
}
[data-testid="stSidebar"] > div {
    padding-top: 1.2rem;
}

/* ── metric cards ── */
[data-testid="metric-container"] {
    background: #13131f;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 12px;
    padding: 16px 20px;
    transition: border-color 0.2s;
}
[data-testid="metric-container"]:hover {
    border-color: rgba(137,212,245,0.25);
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: #6b7280 !important;
    font-size: 0.7rem;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #f0f0f8 !important;
    font-size: 1.5rem;
    font-weight: 700;
    letter-spacing: -0.02em;
}

/* ── tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: #13131f;
    border-radius: 10px;
    padding: 4px;
    gap: 2px;
    border: 1px solid rgba(255,255,255,0.06);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    color: #6b7280 !important;
    font-weight: 500;
    font-size: 0.875rem;
    letter-spacing: 0.01em;
    padding: 8px 22px;
    transition: all 0.15s;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #d1d5db !important;
    background: rgba(255,255,255,0.04) !important;
}
.stTabs [aria-selected="true"] {
    background: rgba(137,212,245,0.08) !important;
    color: #89d4f5 !important;
    border-bottom: 2px solid #89d4f5 !important;
    font-weight: 600 !important;
}

/* ── all buttons ── */
.stButton button {
    background: #13131f;
    border: 1px solid rgba(255,255,255,0.1);
    color: #d1d5db !important;
    border-radius: 8px;
    font-weight: 500;
    font-size: 0.875rem;
    padding: 8px 18px;
    transition: all 0.15s;
    letter-spacing: 0.01em;
}
.stButton button:hover {
    background: #1a1a2e;
    border-color: rgba(137,212,245,0.3);
    color: #89d4f5 !important;
}
.stButton button[kind="primary"] {
    background: rgba(137,212,245,0.1);
    border-color: rgba(137,212,245,0.35);
    color: #89d4f5 !important;
}

/* ── download buttons ── */
[data-testid="stDownloadButton"] button {
    background: rgba(137,212,245,0.08);
    border: 1px solid rgba(137,212,245,0.25);
    color: #89d4f5 !important;
    border-radius: 8px;
    font-weight: 600;
    font-size: 0.875rem;
    letter-spacing: 0.02em;
    transition: all 0.15s;
    width: 100%;
}
[data-testid="stDownloadButton"] button:hover {
    background: rgba(137,212,245,0.16);
    border-color: rgba(137,212,245,0.5);
    box-shadow: 0 0 20px rgba(137,212,245,0.12);
}

/* ── expanders ── */
[data-testid="stExpander"] {
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 12px !important;
    background: #13131f !important;
    transition: border-color 0.2s;
}
[data-testid="stExpander"]:hover {
    border-color: rgba(137,212,245,0.2) !important;
}

/* ── multiselect ── */
[data-baseweb="select"] > div {
    background: #13131f !important;
    border-color: rgba(255,255,255,0.1) !important;
    border-radius: 8px !important;
}
[data-baseweb="tag"] {
    background: rgba(137,212,245,0.1) !important;
    border: 1px solid rgba(137,212,245,0.25) !important;
    color: #89d4f5 !important;
    border-radius: 6px !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
}

/* ── section headings ── */
.stMarkdown h2 {
    color: #f0f0f8;
    font-weight: 700;
    font-size: 1.5rem;
    letter-spacing: -0.02em;
}
.stMarkdown h3 {
    color: #e0e0ec;
    font-weight: 600;
    font-size: 1.1rem;
    letter-spacing: -0.01em;
    border-bottom: 1px solid rgba(255,255,255,0.06);
    padding-bottom: 8px;
    margin-bottom: 4px;
}

/* ── dataframe ── */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 10px;
    overflow: hidden;
}

/* ── number inputs ── */
[data-testid="stNumberInput"] input {
    background: #13131f !important;
    border-color: rgba(255,255,255,0.1) !important;
    border-radius: 8px !important;
    color: #f0f0f8 !important;
    font-family: 'Inter', sans-serif !important;
}
[data-testid="stNumberInput"] input:focus {
    border-color: rgba(137,212,245,0.4) !important;
    box-shadow: 0 0 0 2px rgba(137,212,245,0.08) !important;
}

/* ── text inputs ── */
[data-testid="stTextInput"] input {
    background: #13131f !important;
    border-color: rgba(255,255,255,0.1) !important;
    border-radius: 8px !important;
    color: #f0f0f8 !important;
    font-family: 'Inter', sans-serif !important;
}
[data-testid="stTextInput"] input:focus {
    border-color: rgba(137,212,245,0.4) !important;
    box-shadow: 0 0 0 2px rgba(137,212,245,0.08) !important;
}

/* ── date input ── */
[data-testid="stDateInput"] input {
    background: #13131f !important;
    border-color: rgba(255,255,255,0.1) !important;
    border-radius: 8px !important;
    color: #f0f0f8 !important;
}

/* ── sliders ── */
[data-testid="stSlider"] [data-baseweb="slider"] div[role="slider"] {
    background: #89d4f5 !important;
    border: 2px solid #89d4f5 !important;
    box-shadow: 0 0 8px rgba(137,212,245,0.4) !important;
}
[data-testid="stSlider"] [data-baseweb="slider"] div[data-testid="stSliderTrack"] {
    background: rgba(255,255,255,0.08) !important;
}

/* ── radio buttons ── */
[data-testid="stRadio"] label {
    background: #13131f;
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 8px;
    padding: 6px 14px;
    font-size: 0.85rem;
    color: #9ca3af;
    transition: all 0.15s;
    cursor: pointer;
}
[data-testid="stRadio"] label:hover {
    border-color: rgba(137,212,245,0.25);
    color: #d1d5db;
}
[data-testid="stRadio"] [aria-checked="true"] + div label,
[data-testid="stRadio"] input:checked ~ label {
    border-color: rgba(137,212,245,0.4) !important;
    color: #89d4f5 !important;
    background: rgba(137,212,245,0.06) !important;
}

/* ── selectbox ── */
[data-testid="stSelectbox"] > div > div {
    background: #13131f !important;
    border-color: rgba(255,255,255,0.1) !important;
    border-radius: 8px !important;
}

/* ── checkbox ── */
[data-testid="stCheckbox"] label {
    color: #9ca3af;
    font-size: 0.875rem;
}

/* ── captions ── */
[data-testid="stCaptionContainer"] {
    color: #6b7280 !important;
    font-size: 0.78rem !important;
}

/* ── info/success/warning boxes ── */
[data-testid="stAlert"] {
    border-radius: 10px !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    font-size: 0.875rem;
}

/* ── dividers ── */
hr {
    border: none !important;
    border-top: 1px solid rgba(255,255,255,0.06) !important;
    margin: 1.2rem 0 !important;
}

/* ── sidebar labels ── */
[data-testid="stSidebar"] label {
    color: #9ca3af !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.02em;
}

/* ── data editor ── */
[data-testid="stDataEditor"] {
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 10px !important;
    overflow: hidden;
}
</style>
""", unsafe_allow_html=True)

# ── supabase auth ─────────────────────────────────────────────
from supabase import create_client

@st.cache_resource
def _init_sb():
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

_sb = _init_sb()

def _restore_session():
    tok = st.session_state.get('sb_access_token')
    ref = st.session_state.get('sb_refresh_token')
    if tok and ref:
        try:
            _sb.auth.set_session(tok, ref)
            st.session_state.sb_user = _sb.auth.get_user().user
        except Exception:
            for k in ('sb_access_token','sb_refresh_token','sb_user'):
                st.session_state.pop(k, None)

_restore_session()

if 'sb_user' not in st.session_state or st.session_state.sb_user is None:
    st.markdown("""
    <style>
    .login-wrap {
        max-width: 400px; margin: 80px auto; padding: 40px;
        background: #12121e; border: 1px solid rgba(137,212,245,0.2);
        border-radius: 16px; box-shadow: 0 0 40px rgba(137,212,245,0.08);
    }
    .login-title {
        background: linear-gradient(135deg,#89d4f5,#c9a6ff,#f5a6d3);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-size: 1.6rem; font-weight: 700; letter-spacing: 0.1em; text-align: center;
        margin-bottom: 6px;
    }
    .login-sub { color: #666; text-align: center; font-size: 0.82rem;
                 letter-spacing: 0.06em; margin-bottom: 28px; }
    </style>
    """, unsafe_allow_html=True)

    _lc1, _lc2, _lc3 = st.columns([1, 2, 1])
    with _lc2:
        try:
            st.image("opal_logo.png", use_container_width=True)
        except Exception:
            pass
        st.markdown("<div class='login-title'>FILLRATE</div>", unsafe_allow_html=True)
        st.markdown("<div class='login-sub'>Sign in to continue</div>", unsafe_allow_html=True)

        # pre-fill saved email from localStorage
        import streamlit.components.v1 as _comp_login
        _comp_login.html("""<script>
        try {
          var e = localStorage.getItem('opal_email');
          if (e) {
            var inputs = window.parent.document.querySelectorAll('input[type=text]');
            if (inputs.length > 0) { inputs[0].value = e;
              inputs[0].dispatchEvent(new Event('input', {bubbles:true})); }
          }
        } catch(ex) {}
        </script>""", height=0)

        _mode = st.radio("", ["Sign In", "Create Account"], horizontal=True, label_visibility="collapsed")
        _saved_email = st.session_state.get('_remembered_email', '')
        _email = st.text_input("Email", value=_saved_email, placeholder="you@example.com")
        _pw    = st.text_input("Password", type="password", placeholder="••••••••")
        _remember = st.checkbox("Remember my email", value=bool(_saved_email))

        if _mode == "Create Account":
            _pw2 = st.text_input("Confirm Password", type="password", placeholder="••••••••")
            if st.button("Create Account", use_container_width=True):
                if _pw != _pw2:
                    st.error("Passwords don't match.")
                elif len(_pw) < 6:
                    st.error("Password must be at least 6 characters.")
                else:
                    try:
                        _r = _sb.auth.sign_up({"email": _email, "password": _pw})
                        if _r.user:
                            st.success("Account created! Sign in below.")
                        else:
                            st.error("Sign-up failed. Try again.")
                    except Exception as _e:
                        st.error(str(_e))
        else:
            if st.button("Sign In", use_container_width=True):
                try:
                    _r = _sb.auth.sign_in_with_password({"email": _email, "password": _pw})
                    st.session_state.sb_user          = _r.user
                    st.session_state.sb_access_token  = _r.session.access_token
                    st.session_state.sb_refresh_token = _r.session.refresh_token
                    if _remember:
                        st.session_state['_remembered_email'] = _email
                        _comp_login.html(f"""<script>
                        try {{ localStorage.setItem('opal_email', {repr(_email)}); }} catch(ex) {{}}
                        </script>""", height=0)
                    else:
                        st.session_state.pop('_remembered_email', None)
                        _comp_login.html("""<script>
                        try {{ localStorage.removeItem('opal_email'); }} catch(ex) {{}}
                        </script>""", height=0)
                    st.rerun()
                except Exception as _e:
                    st.error(f"Login error: {str(_e)}")
    st.stop()

# ── stripe ────────────────────────────────────────────────────
import stripe as _stripe
from datetime import datetime as _dt

_stripe.api_key    = st.secrets.get("STRIPE_SECRET_KEY", "")
_PRICE_MONTHLY     = "price_1TVmB3IywzMN4ZEtMUg1VK62"
_PRICE_YEARLY      = "price_1TVmCMIywzMN4ZEtIhU4Ykz5"
_TRIAL_DAYS        = 14
_APP_URL           = st.secrets.get("APP_URL", "https://munchies-order-app.streamlit.app")

def _get_or_create_customer(user):
    res = _sb.table('user_subscriptions').select('stripe_customer_id,status').eq('user_id', user.id).execute()
    if res.data and res.data[0].get('stripe_customer_id'):
        return res.data[0]['stripe_customer_id'], res.data[0].get('status', 'none')
    cust = _stripe.Customer.create(email=user.email, metadata={'uid': user.id})
    _sb.table('user_subscriptions').upsert({'user_id': user.id, 'stripe_customer_id': cust.id, 'status': 'none'}).execute()
    return cust.id, 'none'

def _check_sub(user_id):
    res = _sb.table('user_subscriptions').select('*').eq('user_id', user_id).execute()
    return res.data[0] if res.data else None

def _sf(obj, key, default=None):
    """Safe field access for stripe-python 7.x objects (no .get() method)."""
    try:
        return obj[key]
    except (KeyError, TypeError, AttributeError):
        return default

def _sub_to_row(user_id, cid, sub):
    te = _sf(sub, 'trial_end')
    pe = _sf(sub, 'current_period_end')
    plan = 'yearly' if _sf(sub['items']['data'][0], 'price', {})['id'] == _PRICE_YEARLY else 'monthly'
    return {
        'user_id': user_id,
        'stripe_customer_id': cid,
        'stripe_subscription_id': sub['id'],
        'status': sub['status'],
        'plan': plan,
        'trial_end': _dt.utcfromtimestamp(te).isoformat() if te else None,
        'current_period_end': _dt.utcfromtimestamp(pe).isoformat() if pe else None,
    }

def _sync_sub(user_id, session_id):
    try:
        sess = _stripe.checkout.Session.retrieve(session_id, expand=['subscription'])
        sub  = sess.subscription
        if sub:
            cid = _sf(sess, 'customer') or ''
            _sb.table('user_subscriptions').upsert(_sub_to_row(user_id, cid, sub)).execute()
    except Exception as _e:
        st.error(f"Could not verify payment: {_e}")

def _make_checkout(customer_id, price_id):
    sess = _stripe.checkout.Session.create(
        customer=customer_id,
        payment_method_types=['card'],
        line_items=[{'price': price_id, 'quantity': 1}],
        mode='subscription',
        subscription_data={'trial_period_days': _TRIAL_DAYS},
        allow_promotion_codes=True,
        success_url=f"{_APP_URL}?checkout=success&sid={{CHECKOUT_SESSION_ID}}",
        cancel_url=f"{_APP_URL}?checkout=cancelled",
    )
    return sess.url

# ── handle checkout return ────────────────────────────────────
_qp = st.query_params
if _qp.get('checkout') == 'success' and 'sid' in _qp:
    _sync_sub(st.session_state.sb_user.id, _qp['sid'])
    st.query_params.clear()
    st.rerun()
elif _qp.get('checkout') == 'cancelled':
    st.query_params.clear()
    st.warning("Checkout cancelled — you can subscribe anytime from this page.")

# ── subscription gate ─────────────────────────────────────────
def _sync_from_stripe(user_id, debug=False):
    """Check Stripe directly and sync to Supabase if active subscription found."""
    try:
        res = _sb.table('user_subscriptions').select('stripe_customer_id').eq('user_id', user_id).execute()
        if not res.data or not res.data[0].get('stripe_customer_id'):
            if debug:
                st.info("No Stripe customer linked yet — complete checkout above first.")
            return False
        cid  = res.data[0]['stripe_customer_id']
        subs = _stripe.Subscription.list(customer=cid, limit=10)
        if not subs.data:
            # Try all customers with this email in case a duplicate was created
            user_email = st.session_state.sb_user.email
            customers  = _stripe.Customer.list(email=user_email, limit=10)
            for c in customers.data:
                subs2 = _stripe.Subscription.list(customer=c.id, limit=5)
                if subs2.data:
                    cid  = c.id
                    subs = subs2
                    break
        if subs.data:
            # Pick active/trialing first, else most recent
            sub = next((s for s in subs.data if s['status'] in ('active','trialing')), subs.data[0])
            _sb.table('user_subscriptions').upsert(_sub_to_row(user_id, cid, sub)).execute()
            return sub['status'] in ('active', 'trialing')
        else:
            if debug:
                st.info("No subscription found in Stripe yet — wait a moment after completing checkout, then try again.")
    except Exception as _e:
        if debug:
            import traceback
            st.error(f"Stripe check error ({type(_e).__name__}): {_e}")
            st.code(traceback.format_exc())
    return False

_sub       = _check_sub(st.session_state.sb_user.id)
_is_active = _sub and _sub.get('status') in ('active', 'trialing')

# If not active in Supabase, check Stripe directly (catches post-payment returns)
if not _is_active:
    _is_active = _sync_from_stripe(st.session_state.sb_user.id)
    if _is_active:
        st.rerun()

if not _is_active:
    st.markdown("""
    <style>
    .price-card {
        background: #12121e; border: 1px solid rgba(137,212,245,0.2);
        border-radius: 16px; padding: 32px 24px; text-align: center;
        transition: all 0.2s;
    }
    .price-card:hover { border-color: #89d4f5; box-shadow: 0 0 30px rgba(137,212,245,0.12); }
    .price-badge {
        background: linear-gradient(135deg,rgba(137,212,245,0.15),rgba(201,166,255,0.15));
        border: 1px solid rgba(137,212,245,0.3); border-radius: 20px;
        padding: 4px 14px; font-size: 0.75rem; letter-spacing: 0.08em;
        color: #89d4f5; display: inline-block; margin-bottom: 16px;
    }
    .price-amount { font-size: 2.8rem; font-weight: 800; color: #e8e8f0; }
    .price-period { color: #666; font-size: 0.85rem; margin-top: -4px; }
    .price-save   { color: #89d4f5; font-size: 0.8rem; margin-top: 8px; font-weight: 600; }
    </style>
    """, unsafe_allow_html=True)

    _sc1, _sc2, _sc3 = st.columns([1, 2, 1])
    with _sc2:
        try:
            st.image("opal_logo.png", use_container_width=True)
        except Exception:
            pass
        st.markdown("<div style='text-align:center;margin-bottom:8px'><span style='background:linear-gradient(135deg,#89d4f5,#c9a6ff);-webkit-background-clip:text;-webkit-text-fill-color:transparent;font-size:1.4rem;font-weight:700;letter-spacing:0.1em'>OPAL ORDER TOOL</span></div>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center;color:#666;font-size:0.85rem;margin-bottom:24px'>The smart ordering tool for cannabis retailers</p>", unsafe_allow_html=True)
        st.markdown(f"<div style='text-align:center;background:rgba(137,212,245,0.08);border:1px solid rgba(137,212,245,0.2);border-radius:10px;padding:12px;margin-bottom:24px;color:#89d4f5;font-weight:600'>✨ {_TRIAL_DAYS}-day free trial — no credit card charged until trial ends</div>", unsafe_allow_html=True)

        # Pre-generate checkout URLs once per page load
        if 'checkout_url_monthly' not in st.session_state or 'checkout_url_yearly' not in st.session_state:
            with st.spinner("Preparing checkout..."):
                try:
                    _cid, _ = _get_or_create_customer(st.session_state.sb_user)
                    st.session_state['checkout_url_monthly'] = _make_checkout(_cid, _PRICE_MONTHLY)
                    st.session_state['checkout_url_yearly']  = _make_checkout(_cid, _PRICE_YEARLY)
                except Exception as _ce:
                    st.error(f"Could not connect to payment system: {_ce}")
                    st.stop()

        _url_m = st.session_state['checkout_url_monthly']
        _url_y = st.session_state['checkout_url_yearly']

        import streamlit.components.v1 as _components
        _components.html(f"""
        <style>
        .opal-btn {{
            display:block; width:100%; padding:13px 0;
            background:linear-gradient(135deg,rgba(137,212,245,0.18),rgba(201,166,255,0.18));
            border:1px solid rgba(137,212,245,0.45); border-radius:8px;
            color:#89d4f5; font-family:sans-serif; font-size:0.92rem;
            font-weight:600; letter-spacing:0.04em; text-align:center;
            text-decoration:none; cursor:pointer; margin-bottom:6px;
            transition:all 0.2s;
        }}
        .opal-btn:hover {{
            background:linear-gradient(135deg,rgba(137,212,245,0.32),rgba(201,166,255,0.32));
            border-color:#89d4f5; box-shadow:0 0 16px rgba(137,212,245,0.25);
        }}
        .cards {{ display:flex; gap:16px; }}
        .card {{
            flex:1; background:#12121e; border:1px solid rgba(137,212,245,0.2);
            border-radius:16px; padding:28px 20px; text-align:center;
            font-family:sans-serif;
        }}
        .card.featured {{ border-color:rgba(137,212,245,0.45); }}
        .badge {{
            background:linear-gradient(135deg,rgba(137,212,245,0.15),rgba(201,166,255,0.15));
            border:1px solid rgba(137,212,245,0.3); border-radius:20px;
            padding:4px 14px; font-size:0.72rem; letter-spacing:0.08em;
            color:#89d4f5; display:inline-block; margin-bottom:14px;
        }}
        .amount {{ font-size:2.6rem; font-weight:800; color:#e8e8f0; }}
        .period {{ color:#666; font-size:0.82rem; margin-top:-2px; }}
        .save   {{ color:#89d4f5; font-size:0.78rem; margin-top:6px; font-weight:600; }}
        </style>
        <div class="cards">
          <div class="card">
            <div class="badge">MONTHLY</div>
            <div class="amount">$100</div>
            <div class="period">per month</div>
            <br>
            <a class="opal-btn" href="{_url_m}" target="_blank">Start Free Trial — Monthly</a>
          </div>
          <div class="card featured">
            <div class="badge">YEARLY — BEST VALUE</div>
            <div class="amount">$700</div>
            <div class="period">per year &nbsp;·&nbsp; $58/mo</div>
            <div class="save">Save $500 vs monthly</div>
            <br>
            <a class="opal-btn" href="{_url_y}" target="_blank">Start Free Trial — Yearly</a>
          </div>
        </div>
        """, height=260)

        st.markdown("<p style='text-align:center;color:#555;font-size:0.78rem;margin-top:16px'>Cancel anytime · Secure payments via Stripe · Have a referral code? Enter it at checkout.</p>", unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("✅ I've completed payment — click to activate", use_container_width=True):
            for _k in ('checkout_url_monthly','checkout_url_yearly'):
                st.session_state.pop(_k, None)
            if _sync_from_stripe(st.session_state.sb_user.id, debug=True):
                st.rerun()

        st.markdown("---")
        if st.button("Sign Out", use_container_width=True):
            _sb.auth.sign_out()
            for _k in ('sb_user','sb_access_token','sb_refresh_token'):
                st.session_state.pop(_k, None)
            st.rerun()
    st.stop()

# ── advanced mode definitions ─────────────────────────────────
_SUBCAT_DEFS = [  # kept for profile save compat
    ('Vapes',        ['Closed Loop Pods','Disposable Pens','510 Thread Cartridges'],                    21),
    ('Pre-Roll',     ['Singles','Multipacks'],                                                           14),
    ('Edibles',      ['Baked Goods','Chocolates','Soft Chews','Hard Chews','Confections'],               14),
    ('Topicals',     ['Creams And Lotions','Body Care','Face Care','Patches','Lip Care','Bath'],         30),
    ('Concentrates', ['Shatter','Resin','Rosin','Wax','Distillate','Diamonds'],                         14),
]

_ADV_CATS = ['Flower','Vapes','Pre-Roll','Edibles','Topicals','Concentrates','Oil','Capsules','Beverages','Seeds']

_CAT_SIZE_DEFS = {
    'Flower':       [('1g',7),('2g',14),('3.5g',14),('5g',14),('7g',14),('10g',21),('14g',21),('28g',30),('other',14)],
    'Vapes':        [
        ('0.2g',14),('0.3g',14),('0.4g',14),('0.45g',14),('0.5g',21),
        ('0.95g',21),('1g',21),('1.1g',21),('1.2g',21),('1.25g',21),
        ('1.8g',21),('2g',21),('2x0.5g',21),('3g',21),('4g',21),('4.75g',21),
    ],
    'Pre-Roll':     [
        # Singles
        ('1g',14),('1x0.3g',14),('1x0.35g',14),('1x0.4g',14),('1x0.5g',14),
        ('1x0.6g',14),('1x0.7g',14),('1x0.75g',14),('1x0.8g',14),('1x1g',14),
        ('1x1.5g',14),('1x2g',14),('1x7g',21),('1x14g',30),
        # 2-packs
        ('2x0.35g',21),('2x0.5g',21),('2x0.8g',21),('2x1g',21),
        # 3-packs
        ('3x0.5g',21),('3x0.75g',21),('3x1g',21),
        # 4-packs
        ('4x0.4g',21),('4x0.5g',21),('4x0.75g',21),('4x1g',21),
        # 5-packs
        ('5x0.3g',21),('5x0.35g',21),('5x0.4g',21),('5x0.5g',21),('5x0.6g',21),('5x1g',21),
        # 6-7-packs
        ('6x0.5g',21),('7x0.3g',21),('7x0.5g',21),('7x0.8g',21),
        # 10-packs
        ('10x0.3g',30),('10x0.35g',30),('10x0.4g',30),('10x0.5g',30),
        ('10x0.7g',30),('10x0.75g',30),('10x1g',30),
        # 12-packs
        ('12x0.5g',30),('12x0.6g',30),
        # Large packs
        ('14x0.5g',30),('20x0.35g',30),('20x0.4g',30),('20x0.5g',30),('20x1g',30),
        ('40x0.5g',30),('60x0.5g',30),('70x0.4g',30),('80x0.35g',30),
    ],
    'Edibles':      [
        # Single-unit packs
        ('1 Pack',14),('2 Pack',14),('3 Pack',14),('4 Pack',14),
        ('5 Pack',14),('6 Pack',14),('8 Pack',14),('10 Pack',14),('10 pack',14),
        ('12 Pack',14),('15 Pack',14),('20 Pack',14),('25 Pack',14),
        ('30 Pack',21),('50 Pack',21),
        # Multi-unit retail packs
        ('2x1 Pack',21),('3x4 Pack',21),('4x1 Pack',21),('5x1 Pack',21),('5x2 Pack',21),
        ('6x1 Pack',21),('10x1 Pack',21),('10x2 Pack',21),('10x5 Pack',21),
        ('12x1 Pack',21),('15x1 Pack',21),
        ('20x1 Pack',21),('30x1 Pack',30),('50x1 Pack',30),
    ],
    'Topicals':     [
        ('15ml',21),('20g',21),('25g',21),('25ml',21),
        ('30ml',21),('30g',21),('50ml',30),('50g',30),
        ('60ml',30),('75ml',30),('85g',30),('100ml',30),
        ('125g',30),('130g',30),('250g',30),('250ml',30),('320g',30),
    ],
    'Concentrates': [
        ('0.5g',14),('0.95g',14),('1g',14),('1.2g',14),('1.25g',14),
        ('1.3g',14),('1.5g',14),('2g',14),('2.38g',14),('3.5g',14),
        ('1x0.5g',14),('1x0.7g',14),('1x0.75g',14),('1x0.85g',14),('1x1g',14),
        ('1x1.5g',14),('1x2g',14),('1x3g',21),
        ('2x0.5g',14),('2x0.7g',14),('2x1g',14),
        ('3x0.4g',14),('3x0.45g',14),('3x0.5g',14),('3x0.6g',14),('3x0.75g',14),('3x1g',14),
        ('4x0.5g',21),('5x0.25g',21),('5x0.4g',21),('5x0.5g',21),
        ('6x0.25g',21),('6x0.5g',21),('8x0.25g',21),('8x0.35g',21),
        ('10x0.5g',30),('16x0.25g',30),
    ],
    'Oil':          [
        ('2ml',14),('2.3ml',14),('3ml',14),('3.5ml',14),('5.5g',14),
        ('7ml',14),('7.3g',14),('8.23ml',14),('9ml',14),('9.2ml',14),
        ('10x1ml',14),('12ml',21),('16ml',21),('20ml',21),
        ('30ml',21),('31.2ml',21),('35ml',21),('40ml',21),('45ml',21),
        ('50ml',30),('60ml',30),('90ml',30),
    ],
    'Capsules':     [
        ('5 caps',14),('10 caps',14),('15 caps',14),('18 caps',14),
        ('20 caps',14),('25 caps',14),('30 caps',21),('50 caps',21),
        ('55 caps',21),('60 caps',21),('90 caps',30),('100 caps',30),('200 caps',30),
    ],
    'Beverages':    [
        ('1 Pack',14),('4x1 Pack',14),('65ml',14),('222ml',14),('236ml',14),
        ('330ml',14),('350ml',14),('355ml',14),
        ('4x355ml',21),('6x355ml',21),
    ],
    'Seeds':        [('1 Pack',14),('4 Pack',14),('5 Pack',14),('6 Pack',14)],
}

_CAT_SUBCAT_DEFS = {
    'Flower':       ['Dried Flower','Hash and Kief'],
    'Vapes':        ['510 Thread Cartridges','Closed Loop Pods','Disposable Pens'],
    'Pre-Roll':     ['Pre-Rolls'],
    'Edibles':      ['Baked Goods','Chocolates','Hard Edibles','Pantry','Soft Chews','Sublingual Strips'],
    'Topicals':     ['Bath and Shower','Creams and Lotions','Intimacy Oils','Transdermal'],
    'Concentrates': ['Distillates','Isolates','Resin','Rosin','Shatter','Wax'],
    'Oil':          [],
    'Capsules':     [],
    'Beverages':    [],
    'Seeds':        [],
}

_ADV_DEFAULTS = {'Flower':14,'Vapes':21,'Pre-Roll':14,'Edibles':14,'Topicals':30,'Concentrates':14,'Oil':21,'Capsules':21,'Beverages':14,'Seeds':60}

def _build_settings_from_state(prof_name):
    _prov = st.session_state.get('s_province', 'Ontario')
    _prov_rate = PROVINCE_RATES.get(_prov)
    return {
        'name': prof_name,
        'dos_mode': st.session_state.get('dos_mode', 'Normal'),
        'province': _prov,
        'custom_tax_rate': st.session_state.get('s_custom_tax', 13.0) if _prov_rate is None else None,
        'budget': int(st.session_state.get('s_budget', 30000)),
        'shipping_cost': int(st.session_state.get('s_shipping', 0)),
        'ship_in_budget': bool(st.session_state.get('s_ship_in_budget', False)),
        'flower_simple': int(st.session_state.get('f_all', 14)),
        'flower_targets': {
            '1g': int(st.session_state.get('f1g', 7)),
            '3.5g': int(st.session_state.get('f35g', 14)),
            '5g': int(st.session_state.get('f5g', 14)),
            '7g': int(st.session_state.get('f7g', 14)),
            '14g': int(st.session_state.get('f14g', 21)),
            '28g': int(st.session_state.get('f28g', 30)),
            '30g': int(st.session_state.get('f30g', 30)),
            'other': int(st.session_state.get('fother', 14)),
        },
        'category_targets': {
            'Pre-Roll': int(st.session_state.get('t_preroll', 14)),
            'Edibles': int(st.session_state.get('t_edibles', 14)),
            'Vapes': int(st.session_state.get('t_vapes', 21)),
            'Beverages': int(st.session_state.get('t_beverages', 14)),
            'Capsules': int(st.session_state.get('t_capsules', 21)),
            'Concentrates': int(st.session_state.get('t_conc', 14)),
            'Topicals': int(st.session_state.get('t_topicals', 30)),
            'Oil': int(st.session_state.get('t_oil', 21)),
            'Seeds': int(st.session_state.get('t_seeds', 60)),
        },
        'subcat_targets': {
            f"{_c}||{_sc}": int(st.session_state.get(
                f"sc_{_c}_{_sc}".replace(' ','_').replace('-','_').replace('/','_'), _d))
            for _c, _scs, _d in _SUBCAT_DEFS for _sc in _scs
        },
        'cat_adv_modes': {
            _ac: st.session_state.get(f"adv_{_ac.replace('-','_').replace(' ','_')}_mode", 'By Size')
            for _ac in _ADV_CATS
        },
        'cat_size_targets': {
            f"{_ac}||{_sz}": int(st.session_state.get(
                f"sz_{_ac}_{_sz}".replace('-','_').replace(' ','_'), _d))
            for _ac in _ADV_CATS for _sz, _d in _CAT_SIZE_DEFS[_ac]
        },
        'adv_flower_subcats': {
            _asc: int(st.session_state.get(
                f"sc_Flower_{_asc}".replace(' ','_').replace('-','_').replace('/','_'),
                _ADV_DEFAULTS['Flower']))
            for _asc in _CAT_SUBCAT_DEFS['Flower']
        },
    }

# show trial banner if trialing
if _sub and _sub.get('status') == 'trialing' and _sub.get('trial_end'):
    try:
        _days_left = max(0, (_dt.fromisoformat(_sub['trial_end']) - _dt.utcnow()).days)
        st.info(f"✨ Free trial — {_days_left} day{'s' if _days_left != 1 else ''} remaining. Your card won't be charged until the trial ends.")
    except Exception:
        pass

# ── profile helpers ────────────────────────────────────────────
def _load_profiles():
    try:
        uid = st.session_state.sb_user.id
        rows = _sb.table('location_profiles').select('store_name,settings').eq('id', uid).order('store_name').execute()
        return {r['store_name']: r['settings'] for r in rows.data}
    except Exception:
        return {}

def _save_profile(name, settings):
    uid = st.session_state.sb_user.id
    existing = _sb.table('location_profiles').select('store_name').eq('id', uid).eq('store_name', name).execute()
    if existing.data:
        _sb.table('location_profiles').update({'settings': settings}).eq('id', uid).eq('store_name', name).execute()
    else:
        _sb.table('location_profiles').insert({'id': uid, 'store_name': name, 'settings': settings}).execute()

def _delete_profile(name):
    uid = st.session_state.sb_user.id
    _sb.table('location_profiles').delete().eq('id', uid).eq('store_name', name).execute()

# ── styles ────────────────────────────────────────────────────
def _f(color='000000', bold=False, sz=10): return Font(color=color, bold=bold, size=sz)
def _p(color): return PatternFill('solid', fgColor=color)
def _a(h='left', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _b():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL = _p('1F4E79'); CAT_FILL = _p('2E75B6')
HDR_FONT = _f('FFFFFF', bold=True, sz=10); CAT_FONT = _f('FFFFFF', bold=True, sz=10)
BLD_FONT = _f('000000', bold=True, sz=10); BLK_FONT = _f('000000', sz=10)
RED_FILL = _p('FFC7CE'); YEL_FILL = _p('FFEB9C'); GRN_FILL = _p('C6EFCE')
BORDER   = _b()

def hcell(ws, r, c, val, fill=None, font=None, align='center', wrap=True, colspan=0):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill or HDR_FILL
    cell.font = font or HDR_FONT
    cell.alignment = _a(align, wrap=wrap)
    cell.border = BORDER
    if colspan > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+colspan-1)
    return cell

def vcell(ws, r, c, val, font=None, fill=None, fmt=None, align='left', bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font or (_f('000000', bold=True) if bold else BLK_FONT)
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = _a(align)
    cell.border = BORDER
    return cell

# ── province rates (module-level so profile loader can reference it) ──
PROVINCE_RATES = {
    "Ontario (HST 13%)":              0.13,
    "Nova Scotia (HST 15%)":          0.15,
    "New Brunswick (HST 15%)":        0.15,
    "Newfoundland (HST 15%)":         0.15,
    "PEI (HST 15%)":                  0.15,
    "British Columbia (GST+PST 12%)": 0.12,
    "Manitoba (GST+PST 12%)":         0.12,
    "Saskatchewan (GST+PST 11%)":     0.11,
    "Quebec (GST+QST 14.975%)":       0.14975,
    "Alberta (GST 5%)":               0.05,
    "Custom":                         None,
}
_PROVINCE_LIST = list(PROVINCE_RATES.keys())

# ── shared data loading ────────────────────────────────────────
@st.cache_data(show_spinner="Loading catalogue...")
def load_raw(kova_bytes, ocs_bytes):
    import re
    kova = pd.read_excel(io.BytesIO(kova_bytes), sheet_name='Reorder')
    ocs  = pd.read_excel(io.BytesIO(ocs_bytes),  sheet_name='MasterCatalogue')
    kova_ocs = kova[kova['Supplier'] == 'OCS'].copy()
    kova_ocs['Supplier Sku'] = kova_ocs['Supplier Sku'].str.strip()
    ocs['OCS Variant Number'] = ocs['OCS Variant Number'].str.strip()
    ocs_cols = ['OCS Variant Number','OCS Item Number','Unit Price','Pack Size','Stock Status','Plant Type']
    for _opt in ['Sub-Category','Size','Classification','Category','Product','Product Name','Variant Name','Brand']:
        if _opt in ocs.columns and _opt not in ocs_cols:
            ocs_cols.append(_opt)
    merged = kova_ocs.merge(ocs[ocs_cols], left_on='Supplier Sku', right_on='OCS Variant Number', how='left')
    if 'Sub-Category' not in merged.columns:
        merged['Sub-Category'] = ''
    else:
        merged['Sub-Category'] = merged['Sub-Category'].fillna('')

    def extract_size(sku):
        m = re.search(r'_(\d+\.?\d*[gG])_', str(sku))
        return m.group(1).lower() if m else None

    def extract_preroll_size(sku):
        m = re.search(r'_(\d+x\d+\.?\d*[gG])_', str(sku))
        return m.group(1).lower() if m else None

    def extract_product_size(sku):
        s = str(sku)
        if '___' in s:
            before = s.split('___')[0]
            idx = before.find('_')
            if idx >= 0:
                val = before[idx+1:]
                return val if val else None
        return None

    def map_strain(row):
        pt = str(row.get('Plant Type', '') or '').lower()
        if 'indica' in pt:  return 'Indica'
        if 'sativa' in pt:  return 'Sativa'
        if 'hybrid' in pt:  return 'Hybrid'
        if 'blend'  in pt:  return 'Blend'
        name = str(row.get('Product', '') or '').lower()
        for token in ['- indica', 'indica -', '(indica)']:
            if token in name: return 'Indica'
        for token in ['- sativa', 'sativa -', '(sativa)']:
            if token in name: return 'Sativa'
        for token in ['- hybrid', 'hybrid -', '(hybrid)']:
            if token in name: return 'Hybrid'
        for token in ['- blend', 'blend -', '(blend)']:
            if token in name: return 'Blend'
        return 'Unknown'

    _lr_candidates = ['Last Received Date','Date Last Received','Last Receipt Date',
                      'Last Purchase Date','Last PO Date','Last Order Date','Last Ordered Date']
    _lr_col = next((c for c in _lr_candidates if c in merged.columns), None)
    if _lr_col:
        merged['Last Received Date'] = pd.to_datetime(merged[_lr_col], errors='coerce')
    else:
        merged['Last Received Date'] = pd.NaT

    merged['Flower Size']   = merged.apply(
        lambda r: extract_size(r['Supplier Sku']) if r['Classification'] == 'Flower' else None, axis=1)
    merged['Pre-Roll Size'] = merged.apply(
        lambda r: extract_preroll_size(r['Supplier Sku']) if r['Classification'] == 'Pre-Roll' else None, axis=1)
    merged['Product Size']  = merged['Supplier Sku'].apply(extract_product_size)
    merged['Strain'] = merged.apply(map_strain, axis=1)
    merged['Pack Size']  = pd.to_numeric(merged['Pack Size'],  errors='coerce').fillna(1).astype(int)
    merged['Unit Price'] = pd.to_numeric(merged['Unit Price'], errors='coerce')

    if 'Size' in ocs.columns:
        _sz_raw = ocs['Size'].astype(str).str.strip()
        ocs['Product Size'] = _sz_raw.where(_sz_raw.isin(['', 'nan', 'NaN']) == False, None)
        ocs.loc[ocs['Product Size'].isin(['nan', 'NaN', '']), 'Product Size'] = None
    else:
        ocs['Product Size'] = ocs['OCS Variant Number'].apply(extract_product_size)
    ocs['Strain']       = ocs.apply(lambda r: map_strain({
        'Plant Type': r.get('Plant Type', ''),
        'Product': r.get('Product', r.get('Product Name', r.get('Variant Name', ''))),
    }), axis=1)
    ocs['Pack Size']    = pd.to_numeric(ocs['Pack Size'],  errors='coerce').fillna(1).astype(int)
    ocs['Unit Price']   = pd.to_numeric(ocs['Unit Price'], errors='coerce')
    _SC_TO_CLASS = {
        'Dried Flower': 'Flower', 'Hash and Kief': 'Flower',
        'Pre-Rolls': 'Pre-Roll',
        '510 Thread Cartridges': 'Vapes', 'Closed Loop Pods': 'Vapes', 'Disposable Pens': 'Vapes',
        'Baked Goods': 'Edibles', 'Chocolates': 'Edibles', 'Hard Edibles': 'Edibles',
        'Soft Chews': 'Edibles', 'Pantry': 'Edibles', 'Sublingual Strips': 'Edibles',
        'Beverages': 'Beverages',
        'Distillates': 'Concentrates', 'Isolates': 'Concentrates', 'Resin': 'Concentrates',
        'Rosin': 'Concentrates', 'Shatter': 'Concentrates', 'Wax': 'Concentrates',
        'Oils': 'Oil',
        'Capsules': 'Capsules',
        'Bath and Shower': 'Topicals', 'Creams and Lotions': 'Topicals',
        'Intimacy Oils': 'Topicals', 'Transdermal': 'Topicals',
        'Seeds': 'Seeds',
    }
    if 'Sub-Category' in ocs.columns:
        ocs['Classification'] = ocs['Sub-Category'].map(_SC_TO_CLASS)
    elif 'Category' in ocs.columns:
        ocs['Classification'] = ocs['Category']

    if 'Product' not in ocs.columns:
        for _pn in ['Product Name', 'Variant Name']:
            if _pn in ocs.columns:
                ocs['Product'] = ocs[_pn]
                break
    if 'Product' not in ocs.columns:
        ocs['Product'] = ocs['OCS Variant Number']

    return merged, ocs

# ── sidebar ───────────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("opal_logo.png", use_container_width=True)
    except Exception:
        st.markdown("## 💎 OPAL Order Tool")
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── location profiles ─────────────────────────────────────
    st.markdown("---")
    st.markdown(f"**📍 Location Profiles**")
    st.caption(f"Signed in as {st.session_state.sb_user.email}")

    if st.button("Sign Out", use_container_width=True):
        _sb.auth.sign_out()
        for _k in ('sb_user','sb_access_token','sb_refresh_token'):
            st.session_state.pop(_k, None)
        st.rerun()

    # Load profiles from Supabase
    if 'sb_profiles' not in st.session_state:
        st.session_state.sb_profiles = _load_profiles()

    _profiles = st.session_state.sb_profiles
    _prof_names = list(_profiles.keys())

    def _apply_profile(p):
        if p.get('province') in _PROVINCE_LIST:
            st.session_state['s_province'] = p['province']
        if p.get('custom_tax_rate') is not None:
            st.session_state['s_custom_tax'] = float(p['custom_tax_rate'])
        st.session_state['s_budget']         = int(p.get('budget', 30000))
        st.session_state['s_shipping']       = int(p.get('shipping_cost', 0))
        st.session_state['s_ship_in_budget'] = bool(p.get('ship_in_budget', False))
        st.session_state['prof_name']        = p.get('name', 'My Store')
        if p.get('dos_mode') in ('Simple','Normal','Advanced'):
            st.session_state['dos_mode'] = p['dos_mode']
        if p.get('flower_simple') is not None:
            st.session_state['f_all'] = int(p['flower_simple'])
        _ft = p.get('flower_targets', {})
        for _k, _sk in [('1g','f1g'),('3.5g','f35g'),('5g','f5g'),('7g','f7g'),
                         ('14g','f14g'),('28g','f28g'),('30g','f30g'),('other','fother')]:
            if _k in _ft: st.session_state[_sk] = int(_ft[_k])
        _ct = p.get('category_targets', {})
        for _k, _sk in [('Pre-Roll','t_preroll'),('Edibles','t_edibles'),('Vapes','t_vapes'),
                         ('Beverages','t_beverages'),('Capsules','t_capsules'),
                         ('Concentrates','t_conc'),('Topicals','t_topicals'),
                         ('Oil','t_oil'),('Seeds','t_seeds')]:
            if _k in _ct: st.session_state[_sk] = int(_ct[_k])
        for _encoded, _val in p.get('subcat_targets', {}).items():
            if '||' in _encoded:
                _scat, _ssc = _encoded.split('||', 1)
                _ssk = f"sc_{_scat}_{_ssc}".replace(' ','_').replace('-','_').replace('/','_')
                st.session_state[_ssk] = int(_val)
        for _ac, _am in p.get('cat_adv_modes', {}).items():
            st.session_state[f"adv_{_ac.replace('-','_').replace(' ','_')}_mode"] = _am
        for _encoded, _val in p.get('cat_size_targets', {}).items():
            if '||' in _encoded:
                _ac, _sz = _encoded.split('||', 1)
                st.session_state[f"sz_{_ac}_{_sz}".replace('-','_').replace(' ','_')] = int(_val)
        for _asc, _val in p.get('adv_flower_subcats', {}).items():
            _sk = f"sc_Flower_{_asc}".replace(' ','_').replace('-','_').replace('/','_')
            st.session_state[_sk] = int(_val)

    if _prof_names:
        _sel = st.selectbox("Load a saved location", ["— select —"] + _prof_names, key="prof_select")
        _lc1, _lc2 = st.columns(2)
        if _lc1.button("Load", use_container_width=True) and _sel != "— select —":
            _apply_profile(_profiles[_sel])
            st.session_state['prof_name'] = _sel
            st.rerun()
        if _lc2.button("Delete", use_container_width=True) and _sel != "— select —":
            _delete_profile(_sel)
            st.session_state.sb_profiles.pop(_sel, None)
            st.rerun()

    _prof_name = st.text_input("Store / Location Name", value="My Store", key="prof_name")

    if st.button("💾 Save Profile", use_container_width=True, key="save_prof_top"):
        _s = _build_settings_from_state(_prof_name)
        try:
            _save_profile(_prof_name, _s)
            st.session_state.sb_profiles[_prof_name] = _s
            st.success(f"✅ Saved: {_prof_name}")
        except Exception as _pe:
            st.error(f"Save failed: {_pe}")

    st.markdown("---")
    st.markdown("**📁 Upload Files**")
    kova_file = st.file_uploader("Cova Reorder Report (.xlsx)", type="xlsx", key="kova")
    ocs_file  = st.file_uploader("OCS Catalogue (.xlsx)",       type="xlsx", key="ocs")

    # ── shared data-derived sizes (used by both DoS Advanced and LRC Advanced) ──
    _ALL_LRC_CATS = ['Flower','Vapes','Pre-Roll','Edibles','Concentrates',
                     'Topicals','Oil','Capsules','Beverages','Seeds']
    def _sort_size(s):
        import re as _re
        m = _re.match(r'(\d+\.?\d*)', str(s))
        return (float(m.group(1)) if m else 999, str(s))
    if kova_file and ocs_file:
        _sz_merged, _sz_ocs = load_raw(kova_file.getvalue(), ocs_file.getvalue())
        _shared_cat_sizes = {}
        for _sc in set(_ALL_LRC_CATS + _ADV_CATS):
            _sz_col = 'Flower Size' if _sc == 'Flower' else 'Product Size'
            # Kova history sizes
            _kova_szs = {s for s in _sz_merged[_sz_merged['Classification']==_sc][_sz_col].dropna().unique() if s}
            # Full OCS catalogue sizes (if Classification available in ocs_df)
            _ocs_szs = set()
            if 'Classification' in _sz_ocs.columns and 'Product Size' in _sz_ocs.columns:
                _ocs_szs = {s for s in _sz_ocs[_sz_ocs['Classification']==_sc]['Product Size'].dropna().unique() if s}
            _all_szs = _kova_szs | _ocs_szs
            # Fall back to hardcoded if data has nothing
            if not _all_szs:
                _all_szs = {sz for sz,_ in _CAT_SIZE_DEFS.get(_sc,[])}
            _shared_cat_sizes[_sc] = sorted(_all_szs, key=_sort_size)
    else:
        _shared_cat_sizes = {_sc: [sz for sz,_ in _CAT_SIZE_DEFS.get(_sc,[])]
                             for _sc in set(_ALL_LRC_CATS + _ADV_CATS)}

    st.markdown("---")
    _stab1, _stab2, _stab3 = st.tabs(["⚙️ Budget & Tax", "📅 Days of Supply", "🗓️ Filters"])

    with _stab1:
        province = st.selectbox("Province / Tax Region", _PROVINCE_LIST,
                                index=_PROVINCE_LIST.index(st.session_state.get('s_province', _PROVINCE_LIST[0])),
                                key="s_province")
        if PROVINCE_RATES[province] is None:
            tax_rate = st.number_input("Custom Tax Rate (%)", value=st.session_state.get('s_custom_tax', 13.0),
                                       min_value=0.0, max_value=30.0, step=0.1, key="s_custom_tax") / 100
        else:
            tax_rate = PROVINCE_RATES[province]
            st.caption(f"Tax rate: {tax_rate*100:.3g}%")

        budget = st.number_input("Weekly Budget (tax-in)", value=st.session_state.get('s_budget', 30000),
                                 step=500, format="%d", key="s_budget",
                                 help="Your total weekly spend limit including tax. The app prioritizes the most urgent items to fit within this amount. Items that don't fit are moved to the Deferred list.")

        st.markdown("**Shipping**")
        shipping_cost  = st.number_input("Shipping Cost ($)", value=st.session_state.get('s_shipping', 0),
                                         min_value=0, step=5, format="%d", key="s_shipping",
                                         help="Your OCS delivery fee per order. Leave at 0 if shipping is free or included.")
        ship_in_budget = st.checkbox("Deduct shipping from order budget",
                                      value=st.session_state.get('s_ship_in_budget', False),
                                      key="s_ship_in_budget",
                                      help="When checked, the shipping cost is subtracted from your budget before calculating how much product you can order. Uncheck to track shipping separately.")

        effective_budget = budget - shipping_cost if ship_in_budget else budget
        budget_pretax    = round(effective_budget / (1 + tax_rate), 2)
        tax_amount       = effective_budget - budget_pretax
        st.caption(f"Pre-tax: ${budget_pretax:,.2f}  |  Tax ({tax_rate*100:.3g}%): ${tax_amount:,.2f}"
                   + (f"  |  Shipping deducted: ${shipping_cost:,}" if ship_in_budget and shipping_cost > 0 else ""))

        st.markdown("---")
        _th_title, _th_btn = st.columns([2,1])
        _th_title.markdown("**Tier Thresholds** *(weekly units sold)*")
        if _th_btn.button("Reset", key="reset_tiers", use_container_width=True):
            for _k, _v in [('s_tier_a',7),('s_tier_b',3),('s_tier_c',1),('s_tier_d',0.5)]:
                st.session_state[_k] = _v
            st.rerun()
        st.caption("Items are graded A→D based on weekly velocity. Items below D are excluded from orders.")
        _tc1, _tc2 = st.columns(2)
        tier_a = _tc1.number_input("A — Fast (≥)", value=st.session_state.get('s_tier_a', 7), min_value=1, max_value=99, step=1, key="s_tier_a",
                                    help="Top sellers. These items move fast and should always be in stock. The app orders enough to last your full target days of supply.")
        tier_b = _tc2.number_input("B — Mid (≥)",  value=st.session_state.get('s_tier_b', 3), min_value=1, max_value=99, step=1, key="s_tier_b",
                                    help="Solid movers. The app orders enough to keep about 10 days of stock — not as aggressive as A tier but still a priority.")
        _tc3, _tc4 = st.columns(2)
        tier_c = _tc3.number_input("C — Slow (≥)", value=st.session_state.get('s_tier_c', 1), min_value=1, max_value=99, step=1, key="s_tier_c",
                                    help="Slow movers. The app orders 1 case at a time just to keep the shelf from going empty.")
        tier_d = _tc4.number_input("D — Very Slow (≥)", value=st.session_state.get('s_tier_d', 0.5), min_value=0.0, max_value=99.0, step=0.5, key="s_tier_d", format="%.1f",
                                    help="Barely moving items. The app orders 1 case only if you're completely out of stock. Set to 0 to include all items regardless of velocity.")

        st.markdown("---")
        st.markdown("**Delivery Dates**")
        _today = date.today()
        delivery_date   = st.date_input("Expected Delivery Date", value=st.session_state.get('s_delivery_date', _today + __import__('datetime').timedelta(days=4)),
                                         min_value=_today, key="s_delivery_date",
                                         help="The date you expect to receive this order from OCS. Used to identify Critical items — products that will sell out before your order even arrives.")
        next_order_date = st.date_input("Next Order Date", value=st.session_state.get('s_next_order_date', _today + __import__('datetime').timedelta(days=7)),
                                         min_value=_today, key="s_next_order_date",
                                         help="The date you plan to place your next order. Used to identify At Risk items — products that will be fine for now but will sell out before your next order arrives.")
        lead_time_days   = max(1, (delivery_date   - _today).days)
        _next_lead       = lead_time_days
        _next_delivery   = next_order_date + __import__('datetime').timedelta(days=_next_lead)
        order_cycle_days = max(1, (_next_delivery - _today).days)
        st.caption(f"Critical = out before {delivery_date.strftime('%b %d')} ({lead_time_days}d) · At Risk = out before {_next_delivery.strftime('%b %d')} ({order_cycle_days}d)")

    with _stab2:
        st.markdown("**Order Calculation**")

        vel_window = st.radio("Velocity Window", ["30d", "7d", "Blended"],
                              index=["30d","7d","Blended"].index(st.session_state.get('s_vel_window','Blended')),
                              horizontal=True, key="s_vel_window",
                              help="Which sales window to use when calculating how fast items sell.\n\n• 30d — stable average, less reactive to recent spikes\n• 7d — reflects the most recent week, more reactive\n• Blended — 60% weight on last 7 days, 40% on 30 days. Best default for most stores.")

        safety_stock = st.number_input("Safety Stock Buffer (days)", value=st.session_state.get('s_safety_stock', 3),
                                        min_value=0, max_value=30, step=1, key="s_safety_stock",
                                        help="Extra days of stock added on top of your target to account for pack size rounding and demand variability. If your target is 14 days but you're consistently landing at 11, set this to 3 to close the gap.")

        st.markdown("---")
        st.markdown("**Target Days of Supply**")
        st.caption("How many days of stock to aim for after ordering. Higher = bigger orders but more cash tied up in inventory.")
        _dos_mode = st.radio("Mode", ["Simple", "Normal", "Advanced"],
                             index=["Simple","Normal","Advanced"].index(st.session_state.get('dos_mode','Normal')),
                             horizontal=True, key="dos_mode",
                             help="Simple: one target for all categories. Normal: set targets per category. Advanced: set targets by size or subcategory within each category.")

        def _ni(label, default_key, default_val, col=None):
            _w = col if col else st
            return _w.number_input(label, value=st.session_state.get(default_key, default_val),
                                   min_value=1, max_value=90, key=default_key)

        # Defaults pulled from session_state (overridden by inputs below)
        t_preroll   = st.session_state.get('t_preroll',   14)
        t_edibles   = st.session_state.get('t_edibles',   14)
        t_vapes     = st.session_state.get('t_vapes',     21)
        t_beverages = st.session_state.get('t_beverages', 14)
        t_capsules  = st.session_state.get('t_capsules',  21)
        t_conc      = st.session_state.get('t_conc',      14)
        t_topicals  = st.session_state.get('t_topicals',  30)
        t_oil       = st.session_state.get('t_oil',       21)
        t_seeds     = st.session_state.get('t_seeds',     60)
        SUBCAT_TARGET   = {}
        CAT_SIZE_TARGET = {}
        CAT_ADV_MODES   = {}

        if _dos_mode == "Simple":
            _dc1, _dc2 = st.columns(2)
            t_flower_all = _ni("Flower",       "f_all",       14, _dc1)
            t_preroll    = _ni("Pre-Roll",     "t_preroll",   14, _dc1)
            t_edibles    = _ni("Edibles",      "t_edibles",   14, _dc1)
            t_vapes      = _ni("Vapes",        "t_vapes",     21, _dc1)
            t_beverages  = _ni("Beverages",    "t_beverages", 14, _dc1)
            t_capsules   = _ni("Capsules",     "t_capsules",  21, _dc2)
            t_conc       = _ni("Concentrates", "t_conc",      14, _dc2)
            t_topicals   = _ni("Topicals",     "t_topicals",  30, _dc2)
            t_oil        = _ni("Oil",          "t_oil",       21, _dc2)
            t_seeds      = _ni("Seeds",        "t_seeds",     60, _dc2)
            FLOWER_SIZE_TARGET = {s: t_flower_all for s in ['1g','3.5g','5g','7g','14g','28g','30g','other']}

        elif _dos_mode == "Normal":
            with st.expander("🌸 Flower (by size)", expanded=True):
                fc1, fc2 = st.columns(2)
                t_flower_1g    = _ni("1g",    "f1g",    7,  fc1)
                t_flower_35g   = _ni("3.5g",  "f35g",  14,  fc1)
                t_flower_5g    = _ni("5g",    "f5g",   14,  fc1)
                t_flower_7g    = _ni("7g",    "f7g",   14,  fc1)
                t_flower_14g   = _ni("14g",   "f14g",  21,  fc2)
                t_flower_28g   = _ni("28g",   "f28g",  30,  fc2)
                t_flower_30g   = _ni("30g",   "f30g",  30,  fc2)
                t_flower_other = _ni("Other", "fother",14,  fc2)
            FLOWER_SIZE_TARGET = {
                '1g': t_flower_1g, '3.5g': t_flower_35g, '5g': t_flower_5g,
                '7g': t_flower_7g, '14g': t_flower_14g, '28g': t_flower_28g,
                '30g': t_flower_30g, 'other': t_flower_other,
            }
            _dc1, _dc2 = st.columns(2)
            t_preroll   = _ni("Pre-Roll",     "t_preroll",   14, _dc1)
            t_edibles   = _ni("Edibles",      "t_edibles",   14, _dc1)
            t_vapes     = _ni("Vapes",        "t_vapes",     21, _dc1)
            t_beverages = _ni("Beverages",    "t_beverages", 14, _dc1)
            t_capsules  = _ni("Capsules",     "t_capsules",  21, _dc1)
            t_conc      = _ni("Concentrates", "t_conc",      14, _dc2)
            t_topicals  = _ni("Topicals",     "t_topicals",  30, _dc2)
            t_oil       = _ni("Oil",          "t_oil",       21, _dc2)
            t_seeds     = _ni("Seeds",        "t_seeds",     60, _dc2)

        else:  # Advanced — per-category size vs subcat toggle
            _sz_defaults = {_acat: dict(_CAT_SIZE_DEFS.get(_acat,[])) for _acat in _ADV_CATS}

            for _acat in _ADV_CATS:
                _akey  = f"adv_{_acat.replace('-','_').replace(' ','_')}_mode"
                _icon  = "🌸" if _acat == "Flower" else "🔹"
                _has_subcats = bool(_CAT_SUBCAT_DEFS.get(_acat))
                with st.expander(f"{_icon} {_acat}"):
                    if _has_subcats:
                        _atog = st.radio("", ["By Size","By Subcategory"], horizontal=True,
                                         index=["By Size","By Subcategory"].index(
                                             st.session_state.get(_akey,"By Size")),
                                         key=_akey, label_visibility="collapsed")
                    else:
                        _atog = "By Size"
                    CAT_ADV_MODES[_acat] = _atog
                    _ac1, _ac2 = st.columns(2)
                    if _atog == "By Size":
                        for _ai, _asz in enumerate(_shared_cat_sizes[_acat]):
                            _adef = _sz_defaults[_acat].get(_asz, _ADV_DEFAULTS.get(_acat, 14))
                            _ask  = f"sz_{_acat}_{_asz}".replace('-','_').replace(' ','_')
                            _v    = (_ac1 if _ai%2==0 else _ac2).number_input(
                                _asz, value=st.session_state.get(_ask, _adef),
                                min_value=1, max_value=90, key=_ask)
                            CAT_SIZE_TARGET[(_acat, _asz)] = _v
                    else:
                        _def_sc = _ADV_DEFAULTS.get(_acat, 14)
                        for _ai, _asc in enumerate(_CAT_SUBCAT_DEFS[_acat]):
                            _ask = f"sc_{_acat}_{_asc}".replace(' ','_').replace('-','_').replace('/','_')
                            _v   = (_ac1 if _ai%2==0 else _ac2).number_input(
                                _asc, value=st.session_state.get(_ask, _def_sc),
                                min_value=1, max_value=90, key=_ask)
                            SUBCAT_TARGET[(_acat, _asc)] = _v

            if CAT_ADV_MODES.get('Flower') == 'By Size':
                FLOWER_SIZE_TARGET = {sz: CAT_SIZE_TARGET.get(('Flower', sz), _sz_defaults['Flower'].get(sz, 14))
                                      for sz in _shared_cat_sizes['Flower']}
            else:
                FLOWER_SIZE_TARGET = {'other': _ADV_DEFAULTS['Flower']}

        TARGET = {
            'Pre-Roll': t_preroll, 'Edibles': t_edibles, 'Vapes': t_vapes,
            'Beverages': t_beverages, 'Capsules': t_capsules, 'Concentrates': t_conc,
            'Topicals': t_topicals, 'Oil': t_oil, 'Seeds': t_seeds,
            'Flower': st.session_state.get('f_all', 14),
        }

    with _stab3:
        st.markdown("**🗓️ Last Received Cutoff**")
        st.caption("Exclude SKUs not received within N days. 0 = no filter.")

        _lr_mode = st.radio("Cutoff Mode", ["Simple","Advanced"],
                            index=["Simple","Advanced"].index(st.session_state.get('lr_mode','Simple')),
                            horizontal=True, key="lr_mode")

        _lr_cat_defs = {
            'Flower':60,'Vapes':90,'Pre-Roll':60,'Edibles':120,
            'Concentrates':90,'Topicals':180,'Oil':180,
            'Capsules':180,'Beverages':120,'Seeds':365,
        }

        LAST_RECEIVED_CUTOFF = {}

        if _lr_mode == "Simple":
            _lrc1, _lrc2 = st.columns(2)
            for _lri, _lrcat in enumerate(_ALL_LRC_CATS):
                _lrkey = f"lr_{_lrcat.replace('-','_').replace(' ','_')}"
                _lrdef = _lr_cat_defs.get(_lrcat, 90)
                _lrcol = _lrc1 if _lri % 2 == 0 else _lrc2
                _lrv = _lrcol.number_input(
                    _lrcat, value=st.session_state.get(_lrkey, _lrdef),
                    min_value=0, max_value=730, step=30, key=_lrkey,
                    help="0 = include all")
                LAST_RECEIVED_CUTOFF[_lrcat] = _lrv
        else:  # Advanced — per-size per-category
            for _lrcat in _ALL_LRC_CATS:
                _lrcat_key = _lrcat.replace('-','_').replace(' ','_')
                _cat_def   = _lr_cat_defs.get(_lrcat, 90)
                _cat_sizes = _shared_cat_sizes.get(_lrcat, [])
                _icon      = "🌸" if _lrcat == "Flower" else "🔹"
                with st.expander(f"{_icon} {_lrcat}"):
                    _lrkey_cat = f"lra_{_lrcat_key}_all"
                    _lrv_cat = st.number_input(
                        "All sizes (default)", min_value=0, max_value=730, step=30,
                        value=st.session_state.get(_lrkey_cat, _cat_def),
                        key=_lrkey_cat,
                        help="Fallback for sizes not listed below. 0 = no filter.")
                    LAST_RECEIVED_CUTOFF[_lrcat] = _lrv_cat
                    if _cat_sizes:
                        st.caption("Override per size:")
                        _lsc1, _lsc2 = st.columns(2)
                        for _lsi, _lsz in enumerate(_cat_sizes):
                            _lskey = f"lra_{_lrcat_key}_{str(_lsz).replace('.','_').replace(' ','_')}"
                            _lsv   = (_lsc1 if _lsi%2==0 else _lsc2).number_input(
                                str(_lsz), min_value=0, max_value=730, step=30,
                                value=st.session_state.get(_lskey, _lrv_cat),
                                key=_lskey, help="0 = no filter for this size")
                            LAST_RECEIVED_CUTOFF[(_lrcat, _lsz)] = _lsv


# ── main ──────────────────────────────────────────────────────
try:
    st.image("opal_logo.png", width=280)
except Exception:
    st.markdown("# 💎 OPAL Order Tool")
st.markdown(
    f"<p style='color:#666;font-size:0.85rem;letter-spacing:0.08em;margin-top:-8px'>"
    f"ORDER TOOL &nbsp;·&nbsp; {date.today().strftime('%B %d, %Y').upper()}</p>",
    unsafe_allow_html=True
)

if not kova_file or not ocs_file:
    st.info("Upload your **Cova Reorder Report** and **OCS Catalogue** in the sidebar to get started.")
    st.markdown("""
    **How to export from Cova:**
    Reports → Reorder → Export as .xlsx

    **How to export from OCS Wholesale:**
    Catalogue → Download Catalogue → Excel
    """)
    st.stop()

tab1, tab2, tab3, tab4 = st.tabs(["📦 Replenishment Order", "📋 Suggested Order", "🗂️ Menu Builder", "🧠 AI Menu"])


kova_bytes_raw = kova_file.getvalue()
ocs_bytes_raw  = ocs_file.getvalue()
merged_raw, ocs_df = load_raw(kova_bytes_raw, ocs_bytes_raw)

# ── process data ──────────────────────────────────────────────
@st.cache_data(show_spinner="Processing your data...")
def process(kova_bytes, ocs_bytes, target, flower_size_target, budget_pretax,
            subcat_target=None, cat_size_target=None, cat_adv_modes=None,
            last_received_cutoff=None, tier_thresholds=(5, 2, 1, 0.0),
            safety_stock=0, vel_window='30d'):
    merged = load_raw(kova_bytes, ocs_bytes)[0]
    active = merged[merged['Sales (30 Days)'] > 0].copy()

    # Apply last-received date cutoff (keys: "cat" or "cat||size")
    if last_received_cutoff and 'Last Received Date' in active.columns:
        from datetime import datetime as _dtnow
        _today = _dtnow.utcnow().replace(tzinfo=None)
        def _keep_row(row):
            _cat = row['Classification']
            _sz  = str(row.get('Flower Size') if _cat == 'Flower' else row.get('Product Size', '') or '').strip()
            _sz_key = f"{_cat}||{_sz}"
            _days = last_received_cutoff.get(_sz_key, last_received_cutoff.get(_cat, 0))
            if _days == 0: return True
            _lr = row['Last Received Date']
            if pd.isna(_lr): return True
            return (_today - pd.Timestamp(_lr).replace(tzinfo=None)).days <= _days
        active = active[active.apply(_keep_row, axis=1)].copy()
    # ── velocity window ───────────────────────────────────────
    _vel7  = active['Sales (7 Days)']  / 7
    _vel30 = active['Sales (30 Days)'] / 30
    if vel_window == '7d':
        active['Daily Vel'] = _vel7
    elif vel_window == 'blended':
        active['Daily Vel'] = (_vel7 * 0.6 + _vel30 * 0.4)  # weight recent sales more
    else:  # '30d' default
        active['Daily Vel'] = _vel30
    active['Weekly Vel'] = (active['Daily Vel'] * 7).round(2)
    active['Days Left']  = active['Days of Stock Left (30 Days)'].round(1)
    active['Available']  = active['In Stock Qty'] + active['On Order']

    _ta, _tb, _tc, _td = tier_thresholds
    def assign_tier(w):
        if w >= _ta:   return 'A'
        elif w >= _tb: return 'B'
        elif w >= _tc: return 'C'
        elif w >= _td: return 'D'
        else:          return None
    active['Tier'] = active['Weekly Vel'].apply(assign_tier)
    active = active[active['Tier'].notna()].copy()

    def calc_order(row):
        tier = row['Tier']; pack = int(row['Pack Size'])
        avail = row['Available']; dv = row['Daily Vel']
        dl = row['Days Left'] if pd.notna(row['Days Left']) else 0
        ins = row['In Stock Qty']
        cat = row['Classification']
        _sc = row['Sub-Category'] if 'Sub-Category' in row.index and pd.notna(row['Sub-Category']) and row['Sub-Category'] else ''
        _adv_m = (cat_adv_modes or {}).get(cat)
        if cat == 'Flower':
            if _adv_m == 'By Subcategory':
                tgt = (subcat_target or {}).get(('Flower', _sc), flower_size_target.get('other', target.get('Flower', 14)))
            else:
                tgt = flower_size_target.get(row['Flower Size'], flower_size_target.get('other', 14))
        elif _adv_m == 'By Size':
            raw_sz = str(row.get('Product Size', '') or '').strip()
            tgt = (cat_size_target or {}).get((cat, raw_sz), target.get(cat, 14))
        elif _adv_m == 'By Subcategory':
            tgt = (subcat_target or {}).get((cat, _sc), target.get(cat, 14))
        else:
            tgt = target.get(cat, 14)
        if tier == 'A':
            needed = max(0, math.ceil(dv * (tgt + safety_stock)) - avail)
            cases  = math.ceil(needed / pack) if needed > 0 else 0
        elif tier == 'B':
            if dl >= 10 + safety_stock: return 0, 0, tier
            needed = max(0, math.ceil(dv * (10 + safety_stock)) - avail)
            cases  = math.ceil(needed / pack) if needed > 0 else 0
        elif tier == 'C':
            if dl >= 5 + safety_stock and ins > 0: return 0, 0, tier
            cases = 1
        else:
            if ins > 0 or row['On Order'] > 0: return 0, 0, tier
            cases = 1
        return cases, cases * pack, tier

    active[['Cases','Suggest','Tier']] = active.apply(lambda r: pd.Series(calc_order(r)), axis=1)
    active['Suggest']  = active['Suggest'].astype(int)
    active['Cases']    = active['Cases'].astype(int)
    active['Est Cost'] = (active['Suggest'] * active['Unit Price'].fillna(0)).round(2)

    needs = active[active['Suggest'] > 0].copy()
    needs['TP'] = needs['Tier'].map({'A':0,'B':1,'C':2,'D':3})
    needs['DL'] = needs['Days Left'].fillna(0)
    needs = needs.sort_values(['TP','DL']).reset_index(drop=True)

    cum = 0.0; in_b = []; deferred = []
    for _, row in needs.iterrows():
        if cum + row['Est Cost'] <= budget_pretax:
            cum += row['Est Cost']; in_b.append(row)
        else:
            deferred.append(row)

    order_df    = pd.DataFrame(in_b).sort_values(['Classification','DL']).reset_index(drop=True) if in_b else pd.DataFrame()
    deferred_df = pd.DataFrame(deferred).reset_index(drop=True) if deferred else pd.DataFrame()
    all_active  = active.sort_values(['Classification','DL' if 'DL' in active.columns else 'Days Left']).reset_index(drop=True)

    return order_df, deferred_df, all_active

# Serialise LAST_RECEIVED_CUTOFF for cache hashing (tuple keys → "cat||size" strings)
_lrc_serial = {(f"{k[0]}||{k[1]}" if isinstance(k, tuple) else k): v
               for k, v in LAST_RECEIVED_CUTOFF.items()}
_TIER_THRESHOLDS = (tier_a, tier_b, tier_c, tier_d)
_VEL_WINDOW_MAP = {'30d':'30d','7d':'7d','Blended':'blended'}
order_df, deferred_df, all_active = process(kova_bytes_raw, ocs_bytes_raw, TARGET, FLOWER_SIZE_TARGET, budget_pretax, SUBCAT_TARGET, CAT_SIZE_TARGET, CAT_ADV_MODES, _lrc_serial, _TIER_THRESHOLDS, safety_stock, _VEL_WINDOW_MAP[vel_window])

def _add_risk(df, lead_time, order_cycle):
    if df.empty:
        return df
    df = df.copy()
    df['Net Days'] = (df['In Stock Qty'] + df['On Order']) / df['Daily Vel'].replace(0, float('nan'))
    df['Net Days'] = df['Net Days'].fillna(999).round(1)
    def _risk(nd):
        if nd < lead_time:             return 'Critical'
        elif nd < lead_time + order_cycle: return 'At Risk'
        else:                          return 'Stable'
    df['Risk'] = df['Net Days'].apply(_risk)
    return df

order_df    = _add_risk(order_df,    lead_time_days, order_cycle_days)
deferred_df = _add_risk(deferred_df, lead_time_days, order_cycle_days)
all_active  = _add_risk(all_active,  lead_time_days, order_cycle_days)

if order_df.empty:
    st.warning("No items need ordering based on current stock levels and settings.")
    st.stop()

# ── shared style helpers ──────────────────────────────────────
def color_tier(val):
    colors = {'A':'background-color:#C6EFCE','B':'background-color:#E2EFDA',
              'C':'background-color:#FFEB9C','D':'background-color:#FFC7CE'}
    return colors.get(val, '')

def color_days(val):
    try:
        v = float(val)
        if v <= 3:  return 'background-color:#FFC7CE;color:#9C0006;font-weight:bold'
        if v <= 7:  return 'background-color:#FFEB9C;color:#9C6500;font-weight:bold'
        if v >= 14: return 'background-color:#C6EFCE;color:#276221'
    except: pass
    return ''

# ── workbook builders (defined before tabs) ───────────────────
def build_workbook(order_df, deferred_df, all_active, budget_pretax, target, today, tax_rate=0.13):
    wb = Workbook()

    tier_colours = {'A':'375623','B':'375623','C':'9C6500','D':'9C0006'}
    tier_fills   = {'A':'C6EFCE','B':'E2EFDA','C':'FFEB9C','D':'FFC7CE'}

    def write_order_sheet(ws, df, title):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        cols = [
            ('Category',14),('Product Name',40),('OCS Item #',12),('Tier',7),
            ('Days Left',10),('In Stock',10),('On Order',10),('Wkly Sales',10),
            ('30d Sales',10),('Daily Vel.',10),('Pack Size',10),('Cases',8),
            ('Suggest Order',13),('Unit Price',11),('Est. Cost',11),('OCS Variant #',18),
        ]
        for i, (_, w) in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width = w
        hcell(ws, 1, 1, title, colspan=len(cols))
        for i, (h, _) in enumerate(cols, 1): hcell(ws, 2, i, h, wrap=True)
        ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 30

        ROW = 3; prev_cat = None
        for _, row in df.iterrows():
            cat = row['Classification']
            if cat != prev_cat:
                for ci in range(1, len(cols)+1):
                    c = ws.cell(row=ROW, column=ci); c.fill = CAT_FILL; c.border = BORDER
                ws.cell(row=ROW,column=1).value = f'▶  {cat.upper()}'
                ws.cell(row=ROW,column=1).font = CAT_FONT
                ws.cell(row=ROW,column=1).alignment = _a('left')
                ROW += 1; prev_cat = cat

            alt = _p('F5FBFF') if ROW%2==0 else _p('FFFFFF')
            t = row['Tier']
            dl = row['Days Left'] if pd.notna(row.get('Days Left')) else 0

            def v(c, val, fmt=None, align='center', fill=alt):
                vcell(ws, ROW, c, val, fill=fill, fmt=fmt, align=align)

            v(1,  row['Classification'], align='left', fill=alt)
            v(2,  row['Product'],        align='left', fill=alt)
            v(3,  row['OCS Item Number'] if pd.notna(row.get('OCS Item Number')) else row['Supplier Sku'].split('_')[0])
            ct = ws.cell(row=ROW,column=4,value=t)
            ct.fill=_p(tier_fills[t]); ct.font=_f(tier_colours[t],bold=True); ct.alignment=_a('center'); ct.border=BORDER
            dl_c = ws.cell(row=ROW,column=5,value=round(dl,1))
            dl_c.number_format='0.0'; dl_c.alignment=_a('center'); dl_c.border=BORDER
            if dl<=3:   dl_c.fill=RED_FILL; dl_c.font=_f('9C0006',bold=True)
            elif dl<=7: dl_c.fill=YEL_FILL; dl_c.font=_f('9C6500',bold=True)
            else:       dl_c.fill=GRN_FILL; dl_c.font=_f('276221'); dl_c.fill=alt if dl<14 else GRN_FILL
            v(6,  int(row['In Stock Qty']))
            v(7,  int(row['On Order']))
            v(8,  int(row['Sales (7 Days)']))
            v(9,  int(row['Sales (30 Days)']))
            v(10, round(row['Daily Vel'],3), fmt='0.000')
            v(11, int(row['Pack Size']))
            v(12, int(row['Cases']))
            cs = ws.cell(row=ROW,column=13,value=int(row['Suggest']))
            cs.font=_f('000000',bold=True); cs.fill=_p('FFF2CC'); cs.alignment=_a('center'); cs.border=BORDER
            price = row['Unit Price'] if pd.notna(row.get('Unit Price')) else None
            v(14, price, fmt='"$"#,##0.00')
            v(15, row['Est Cost'] if pd.notna(row.get('Est Cost')) and row['Est Cost']>0 else None, fmt='"$"#,##0.00')
            v(16, row['Supplier Sku'], align='left')
            ROW += 1

        last = ROW-1
        for label, offset, formula in [
            ('SUBTOTAL (pre-tax)', 0, f'=SUM(O3:O{last})'),
            (f'Tax ({tax_rate*100:.3g}%)', 1, f'=O{ROW}*{tax_rate}'),
            ('TOTAL (tax-in)',    2, f'=O{ROW}+O{ROW+1}'),
        ]:
            r = ROW+offset
            for ci in range(1,len(cols)+1):
                ws.cell(row=r,column=ci).fill=_p('1F4E79'); ws.cell(row=r,column=ci).border=BORDER
            lbl=ws.cell(row=r,column=1,value=label); lbl.font=HDR_FONT; lbl.fill=_p('1F4E79'); lbl.alignment=_a('right'); lbl.border=BORDER
            cf=ws.cell(row=r,column=15,value=formula); cf.font=HDR_FONT; cf.fill=_p('1F4E79')
            cf.number_format='"$"#,##0.00'; cf.alignment=_a('center'); cf.border=BORDER

    # Sheet 1: Order Builder
    ws_ob = wb.active; ws_ob.title = 'Order Builder'
    write_order_sheet(ws_ob, order_df, f'OCS Order Builder — {today.strftime("%B %d, %Y")} — Budget: ${budget_pretax:,.0f} pre-tax (${round(budget_pretax*(1+tax_rate)):,} tax-in)')
    ws_ob.sheet_properties.tabColor = 'FFC000'

    # Sheet 2: Portal Order
    ws_portal = wb.create_sheet('OCS Portal Order')
    ws_portal.sheet_view.showGridLines = False
    ws_portal.freeze_panes = 'A3'
    pcols = [('OCS Item #',16),('OCS Variant #',22),('Product Name',48),('Category',14),
             ('Cases to Order',14),('Units per Case',13),('Total Units',12),('Unit Price',12),('Line Total',13)]
    for i,(_,w) in enumerate(pcols,1):
        ws_portal.column_dimensions[ws_portal.cell(row=1,column=i).column_letter].width=w
    hcell(ws_portal,1,1,f'OCS Wholesale Order — {today.strftime("%B %d, %Y")}  |  Enter CASES in the portal',colspan=len(pcols))
    for i,(h,_) in enumerate(pcols,1): hcell(ws_portal,2,i,h,wrap=True)
    PROW=3; prev_pcat=None
    portal_sorted = order_df.sort_values(['Classification','Product']).reset_index(drop=True)
    for _,row in portal_sorted.iterrows():
        cat=row['Classification']
        if cat!=prev_pcat:
            for ci in range(1,len(pcols)+1):
                c=ws_portal.cell(row=PROW,column=ci); c.fill=CAT_FILL; c.border=BORDER
            ws_portal.cell(row=PROW,column=1).value=f'▶  {cat.upper()}'
            ws_portal.cell(row=PROW,column=1).font=CAT_FONT; ws_portal.cell(row=PROW,column=1).alignment=_a('left')
            PROW+=1; prev_pcat=cat
        alt=_p('F0F7FF') if PROW%2==0 else _p('FFFFFF')
        item=str(row['OCS Item Number']).split('.')[0] if pd.notna(row.get('OCS Item Number')) else row['Supplier Sku'].split('_')[0]
        ci_=ws_portal.cell(row=PROW,column=1,value=item); ci_.font=_f('000000',bold=True,sz=11); ci_.fill=alt; ci_.border=BORDER; ci_.alignment=_a('center')
        vcell(ws_portal,PROW,2,row['Supplier Sku'],fill=alt,align='left')
        vcell(ws_portal,PROW,3,row['Product'],fill=alt,align='left')
        vcell(ws_portal,PROW,4,cat,fill=alt,align='left')
        cc=ws_portal.cell(row=PROW,column=5,value=int(row['Cases'])); cc.font=_f('000000',bold=True,sz=12); cc.fill=_p('FFF2CC'); cc.alignment=_a('center'); cc.border=BORDER
        vcell(ws_portal,PROW,6,int(row['Pack Size']),fill=alt,align='center')
        vcell(ws_portal,PROW,7,int(row['Suggest']),fill=alt,align='center')
        vcell(ws_portal,PROW,8,row['Unit Price'] if pd.notna(row.get('Unit Price')) else None,fill=alt,fmt='"$"#,##0.00',align='center')
        vcell(ws_portal,PROW,9,row['Est Cost'] if pd.notna(row.get('Est Cost')) and row['Est Cost']>0 else None,fill=alt,fmt='"$"#,##0.00',align='center')
        PROW+=1
    last_p=PROW-1
    for label,offset,formula in [('SUBTOTAL (pre-tax)',0,f'=SUM(I3:I{last_p})'),(f'Tax ({tax_rate*100:.3g}%)',1,f'=I{PROW}*{tax_rate}'),('TOTAL (tax-in)',2,f'=I{PROW}+I{PROW+1}')]:
        r=PROW+offset
        for ci in range(1,len(pcols)+1): ws_portal.cell(row=r,column=ci).fill=_p('1F4E79'); ws_portal.cell(row=r,column=ci).border=BORDER
        lbl=ws_portal.cell(row=r,column=1,value=label); lbl.font=HDR_FONT; lbl.fill=_p('1F4E79'); lbl.alignment=_a('right'); lbl.border=BORDER
        ws_portal.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
        cf=ws_portal.cell(row=r,column=9,value=formula); cf.font=HDR_FONT; cf.fill=_p('1F4E79'); cf.number_format='"$"#,##0.00'; cf.alignment=_a('center'); cf.border=BORDER
    ws_portal.row_dimensions[1].height=22; ws_portal.row_dimensions[2].height=30
    ws_portal.sheet_properties.tabColor='375623'

    # Sheet 3: Deferred
    if not deferred_df.empty:
        ws_def = wb.create_sheet('Deferred (Next Order)')
        write_order_sheet(ws_def, deferred_df, f'Deferred Items — Carry to Next Order (${deferred_df["Est Cost"].sum():,.2f} pre-tax)')
        ws_def.sheet_properties.tabColor='843C0C'

    # Sheet 4: Full Inventory
    ws_inv = wb.create_sheet('Full Inventory')
    ws_inv.sheet_view.showGridLines = False; ws_inv.freeze_panes='A3'
    inv_cols=[('Category',14),('Product Name',40),('OCS Item #',12),('Tier',7),('Days Left',10),
              ('In Stock',10),('On Order',10),('7d Sales',9),('30d Sales',9),('60d Sales',9),
              ('Daily Vel.',10),('Wkly Vel.',10),('Suggest Order',12),('Unit Price',11),('Margin %',10),('OCS Variant #',18)]
    for i,(_,w) in enumerate(inv_cols,1):
        ws_inv.column_dimensions[ws_inv.cell(row=1,column=i).column_letter].width=w
    hcell(ws_inv,1,1,f'Full Active Inventory — {today.strftime("%B %d, %Y")}',colspan=len(inv_cols))
    for i,(h,_) in enumerate(inv_cols,1): hcell(ws_inv,2,i,h,wrap=True)
    ws_inv.row_dimensions[1].height=22; ws_inv.row_dimensions[2].height=30
    IROW=3; prev_icat=None
    for _,row in all_active.iterrows():
        cat=row['Classification']
        if cat!=prev_icat:
            for ci in range(1,len(inv_cols)+1):
                c=ws_inv.cell(row=IROW,column=ci); c.fill=CAT_FILL; c.border=BORDER
            ws_inv.cell(row=IROW,column=1).value=f'▶  {cat.upper()}'
            ws_inv.cell(row=IROW,column=1).font=CAT_FONT; ws_inv.cell(row=IROW,column=1).alignment=_a('left')
            IROW+=1; prev_icat=cat
        t=row['Tier']; alt=_p('F5FBFF') if IROW%2==0 else _p('FFFFFF')
        dl=row['Days Left'] if pd.notna(row.get('Days Left')) else 0
        suggest=int(row['Suggest'])
        def wi(c,val,fmt=None,align='center',fill=alt): vcell(ws_inv,IROW,c,val,fill=fill,fmt=fmt,align=align)
        wi(1,row['Classification'],align='left',fill=alt); wi(2,row['Product'],align='left',fill=alt)
        wi(3,row['OCS Item Number'] if pd.notna(row.get('OCS Item Number')) else row['Supplier Sku'].split('_')[0])
        ct2=ws_inv.cell(row=IROW,column=4,value=t); ct2.fill=_p(tier_fills[t]); ct2.font=_f(tier_colours[t],bold=True); ct2.alignment=_a('center'); ct2.border=BORDER
        dl2=ws_inv.cell(row=IROW,column=5,value=round(dl,1)); dl2.number_format='0.0'; dl2.alignment=_a('center'); dl2.border=BORDER
        if dl<=3: dl2.fill=RED_FILL; dl2.font=_f('9C0006',bold=True)
        elif dl<=7: dl2.fill=YEL_FILL; dl2.font=_f('9C6500',bold=True)
        else: dl2.fill=alt
        wi(6,int(row['In Stock Qty'])); wi(7,int(row['On Order']))
        wi(8,int(row['Sales (7 Days)'])); wi(9,int(row['Sales (30 Days)'])); wi(10,int(row['Sales (60 Days)']))
        wi(11,round(row['Daily Vel'],3),fmt='0.000'); wi(12,round(row['Weekly Vel'],2),fmt='0.00')
        cs2=ws_inv.cell(row=IROW,column=13,value=suggest); cs2.fill=_p('FFF2CC') if suggest>0 else alt; cs2.font=_f('000000',bold=(suggest>0)); cs2.alignment=_a('center'); cs2.border=BORDER
        wi(14,row['Unit Price'] if pd.notna(row.get('Unit Price')) else None,fmt='"$"#,##0.00')
        wi(15,row['Margin'] if pd.notna(row.get('Margin')) else None,fmt='0.0%')
        wi(16,row['Supplier Sku'],align='left')
        IROW+=1
    ws_inv.sheet_properties.tabColor='70AD47'

    wb.active = wb['Order Builder']
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def build_upload(order_df):
    upload = order_df[['Supplier Sku','Cases']].copy()
    upload.columns = ['SKU','Quantity']
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        upload.to_excel(writer, sheet_name='MasterCatalogue', index=False)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════
# TAB 1: REPLENISHMENT ORDER
# ══════════════════════════════════════════════════════════════
with tab1:
    pretax_total      = order_df['Est Cost'].sum()
    tax_amount_actual = pretax_total * tax_rate
    subtotal_taxin    = pretax_total + tax_amount_actual
    grand_total       = subtotal_taxin + (shipping_cost if not ship_in_budget else 0)
    tax_label         = f"Tax ({tax_rate*100:.3g}%)"

    # ── summary metrics ───────────────────────────────────────
    st.markdown("### This Week's Order")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("SKUs to Order",   len(order_df))
    m2.metric("Pre-Tax Total",   f"${pretax_total:,.2f}")
    m3.metric(tax_label,         f"${tax_amount_actual:,.2f}")
    m4.metric("Subtotal (Tax-In)", f"${subtotal_taxin:,.2f}")
    m5.metric("Shipping",        f"${shipping_cost:,}" if shipping_cost > 0 else "—")
    m6.metric("Grand Total",     f"${grand_total:,.2f}")

    st.markdown("---")

    # ── budget breakdown chart ────────────────────────────────
    import plotly.graph_objects as go

    _cat_costs = order_df.groupby('Classification')['Est Cost'].sum().sort_values(ascending=False)
    _tier_costs = order_df.groupby('Tier')['Est Cost'].sum()

    _CAT_COLORS = {
        'Flower':'#89d4f5','Pre-Roll':'#b8e0ff','Vapes':'#7ec8e3',
        'Edibles':'#f5c842','Concentrates':'#e07b39','Beverages':'#6dd6a0',
        'Capsules':'#b89ef5','Oil':'#f5a0c8','Topicals':'#8fd48f','Seeds':'#d4c089',
    }
    _TIER_COLORS = {'A':'#4caf80','B':'#8bc34a','C':'#ffc107','D':'#ef5350'}

    _chart_col1, _chart_col2 = st.columns([3, 2])

    with _chart_col1:
        _fig_cat = go.Figure(go.Bar(
            x=_cat_costs.values,
            y=_cat_costs.index,
            orientation='h',
            marker_color=[_CAT_COLORS.get(c,'#89d4f5') for c in _cat_costs.index],
            text=[f"${v:,.0f}" for v in _cat_costs.values],
            textposition='outside',
            hovertemplate='%{y}: $%{x:,.2f}<extra></extra>',
        ))
        _fig_cat.update_layout(
            title=dict(text="Budget by Category", font=dict(size=13, color='#e8e8f0')),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#e8e8f0', size=11),
            xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
            yaxis=dict(autorange='reversed', tickfont=dict(size=11)),
            margin=dict(l=10, r=60, t=36, b=10),
            height=max(200, len(_cat_costs) * 32 + 60),
        )
        st.plotly_chart(_fig_cat, use_container_width=True, config={'displayModeBar': False})

    with _chart_col2:
        _tier_label_map = {'A':'A — Fast','B':'B — Mid','C':'C — Slow','D':'D — Very Slow'}
        _fig_tier = go.Figure(go.Pie(
            labels=[_tier_label_map.get(t, t) for t in _tier_costs.index],
            values=_tier_costs.values,
            marker_colors=[_TIER_COLORS.get(t,'#aaa') for t in _tier_costs.index],
            hole=0.55,
            textinfo='label+percent',
            hovertemplate='%{label}: $%{value:,.2f}<extra></extra>',
            textfont=dict(size=11),
        ))
        _fig_tier.update_layout(
            title=dict(text="Budget by Tier", font=dict(size=13, color='#e8e8f0')),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#e8e8f0'),
            showlegend=False,
            margin=dict(l=10, r=10, t=36, b=10),
            height=max(200, len(_cat_costs) * 32 + 60),
        )
        st.plotly_chart(_fig_tier, use_container_width=True, config={'displayModeBar': False})

    st.markdown("---")

    # ── tier filter ───────────────────────────────────────────
    st.markdown("### Order Detail")
    tier_options = [f'A — Fast ({tier_a}+/wk)', f'B — Mid ({tier_b}–{tier_a-1}/wk)', f'C — Slow ({tier_c}–{tier_b-1}/wk)', f'D — Very Slow ({tier_d}–{tier_c-1}/wk)']
    tier_map     = {tier_options[0]:'A', tier_options[1]:'B', tier_options[2]:'C', tier_options[3]:'D'}
    selected_labels = st.multiselect(
        "Filter by Tier (select one or more to combine)",
        options=tier_options, default=tier_options, key="t1_tier_filter"
    )
    selected_tiers = [tier_map[l] for l in selected_labels] if selected_labels else list(tier_map.values())
    filtered_df = order_df[order_df['Tier'].isin(selected_tiers)]

    t1_risk = st.radio(
        "Filter by stockout risk (based on Net Days of Stock = In Stock + On Order ÷ Daily Vel)",
        ["All", "At Risk or Critical", "Critical only"],
        horizontal=True, key="t1_risk_filter",
        help=f"Critical = runs out before this order arrives (<{lead_time_days}d) · At Risk = runs out before next order (<{lead_time_days+order_cycle_days}d)"
    )
    if t1_risk == "Critical only":
        filtered_df = filtered_df[filtered_df['Risk'] == 'Critical']
    elif t1_risk == "At Risk or Critical":
        filtered_df = filtered_df[filtered_df['Risk'].isin(['Critical', 'At Risk'])]

    if filtered_df.empty:
        st.info("No items match the selected tiers.")
    else:
        f_pretax    = filtered_df['Est Cost'].sum()
        f_taxin     = f_pretax * (1 + tax_rate)
        f_grand     = f_taxin + (shipping_cost if not ship_in_budget else 0)
        fa1, fa2, fa3, fa4, fa5 = st.columns(5)
        fa1.metric("SKUs (filtered)", len(filtered_df))
        fa2.metric("Pre-Tax",   f"${f_pretax:,.2f}")
        fa3.metric(tax_label,   f"${f_pretax * tax_rate:,.2f}")
        fa4.metric("Shipping",  f"${shipping_cost:,}" if shipping_cost > 0 else "—")
        fa5.metric("Grand Total", f"${f_grand:,.2f}")

        _edit_df = filtered_df[['Classification','Product','Risk','Net Days','Tier','Days Left',
                                 'In Stock Qty','On Order','Sales (7 Days)','Sales (30 Days)',
                                 'Weekly Vel','Daily Vel','Cases','Suggest','Unit Price','Est Cost',
                                 'Pack Size','Supplier Sku']].copy()
        _edit_df['Days Left'] = _edit_df['Days Left'].round(1)
        _edited = st.data_editor(
            _edit_df,
            hide_index=True,
            use_container_width=True,
            height=500,
            column_config={
                'Classification': st.column_config.TextColumn('Category',    disabled=True),
                'Product':        st.column_config.TextColumn('Product',     disabled=True, width='large'),
                'Risk':           st.column_config.TextColumn('Risk',        disabled=True, width='small'),
                'Net Days':       st.column_config.NumberColumn('Net Days',  disabled=True, format='%.1f'),
                'Tier':           st.column_config.TextColumn('Tier',        disabled=True, width='small'),
                'Days Left':      st.column_config.NumberColumn('Days Left', disabled=True, format='%.1f'),
                'In Stock Qty':   st.column_config.NumberColumn('In Stock',  disabled=True),
                'On Order':       st.column_config.NumberColumn('On Order',  disabled=True),
                'Sales (7 Days)': st.column_config.NumberColumn('7d Sales',  disabled=True),
                'Sales (30 Days)':st.column_config.NumberColumn('30d Sales', disabled=True),
                'Weekly Vel':     st.column_config.NumberColumn('Wkly Vel',  disabled=True, format='%.2f'),
                'Daily Vel':      st.column_config.NumberColumn('Daily Vel', disabled=True, format='%.3f'),
                'Cases':          st.column_config.NumberColumn('Cases ✏️',  disabled=False, min_value=0, step=1),
                'Suggest':        st.column_config.NumberColumn('Units',     disabled=True),
                'Unit Price':     st.column_config.NumberColumn('Unit Price',disabled=True, format='$%.2f'),
                'Est Cost':       st.column_config.NumberColumn('Est Cost',  disabled=True, format='$%.2f'),
                'Pack Size':      st.column_config.NumberColumn('Pack Size', disabled=True),
                'Supplier Sku':   st.column_config.TextColumn('OCS Variant #', disabled=True),
            },
            key="order_editor",
        )
        # Recalculate Units and Est Cost from any edited Cases
        _edited['Suggest']  = (_edited['Cases'] * _edited['Pack Size']).astype(int)
        _edited['Est Cost'] = (_edited['Suggest'] * _edited['Unit Price'].fillna(0)).round(2)
        filtered_df = _edited  # downstream downloads use the edited version

    if not deferred_df.empty:
        with st.expander(f"📋 Deferred to Next Order ({len(deferred_df)} SKUs — ${deferred_df['Est Cost'].sum():,.2f} pre-tax)"):
            def_show = deferred_df[['Classification','Product','Tier','Days Left','In Stock Qty','Weekly Vel','Cases','Est Cost']].copy()
            def_show['Est Cost'] = def_show['Est Cost'].map(lambda x: f'${x:,.2f}' if pd.notna(x) else '—')
            st.dataframe(def_show.rename(columns={'Classification':'Category','In Stock Qty':'In Stock','Weekly Vel':'Wkly Vel'}),
                         hide_index=True, use_container_width=True)

    st.markdown("---")

    # ── download files ────────────────────────────────────────
    st.markdown("### Download Files")
    if not filtered_df.empty:
        dl1, dl2 = st.columns(2)
        today = date.today()
        with dl1:
            workbook_buf = build_workbook(filtered_df, deferred_df, all_active, budget_pretax, TARGET, today, tax_rate)
            st.download_button(
                label="📥 Download Order Workbook (.xlsx)",
                data=workbook_buf,
                file_name=f'OCS_Order_Tool_{today.strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        with dl2:
            upload_buf = build_upload(filtered_df)
            st.download_button(
                label="📤 Download OCS Upload File (.xlsx)",
                data=upload_buf,
                file_name=f'OCS_Upload_{today.strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )

# ══════════════════════════════════════════════════════════════
# TAB 2: SUGGESTED ORDER (NO BUDGET CAP)
# ══════════════════════════════════════════════════════════════
with tab2:
    st.markdown("### 📋 Suggested Order")
    st.caption("All items that need restocking based on velocity — no hard budget cap. Review the total before submitting.")

    uncapped_df, _, _ = process(kova_bytes_raw, ocs_bytes_raw, TARGET, FLOWER_SIZE_TARGET, 999_999_999, SUBCAT_TARGET, CAT_SIZE_TARGET, CAT_ADV_MODES, _lrc_serial, _TIER_THRESHOLDS, safety_stock, _VEL_WINDOW_MAP[vel_window])
    uncapped_df = _add_risk(uncapped_df, lead_time_days, order_cycle_days)

    if uncapped_df.empty:
        st.success("✅ Nothing needs restocking right now.")
    else:
        # ── tier filter ───────────────────────────────────────
        uc_tier_options = [f'A — Fast ({tier_a}+/wk)', f'B — Mid ({tier_b}–{tier_a-1}/wk)', f'C — Slow ({tier_c}–{tier_b-1}/wk)', f'D — Very Slow ({tier_d}–{tier_c-1}/wk)']
        uc_tier_map     = {uc_tier_options[0]:'A', uc_tier_options[1]:'B', uc_tier_options[2]:'C', uc_tier_options[3]:'D'}
        uc_selected_labels = st.multiselect(
            "Filter by Tier (select one or more to combine)",
            options=uc_tier_options, default=uc_tier_options, key="t2_tier_filter"
        )
        uc_selected_tiers = [uc_tier_map[l] for l in uc_selected_labels] if uc_selected_labels else list(uc_tier_map.values())
        uc_filtered = uncapped_df[uncapped_df['Tier'].isin(uc_selected_tiers)]

        uc_risk = st.radio(
            "Filter by stockout risk (based on Net Days of Stock = In Stock + On Order ÷ Daily Vel)",
            ["All", "At Risk or Critical", "Critical only"],
            horizontal=True, key="t2_risk_filter",
            help=f"Critical = runs out before this order arrives (<{lead_time_days}d) · At Risk = runs out before next order (<{lead_time_days+order_cycle_days}d)"
        )
        if uc_risk == "Critical only":
            uc_filtered = uc_filtered[uc_filtered['Risk'] == 'Critical']
        elif uc_risk == "At Risk or Critical":
            uc_filtered = uc_filtered[uc_filtered['Risk'].isin(['Critical', 'At Risk'])]

        uc_pretax = uc_filtered['Est Cost'].sum()
        uc_tax    = uc_pretax * tax_rate
        uc_taxin  = uc_pretax + uc_tax
        uc_grand  = uc_taxin + (shipping_cost if not ship_in_budget else 0)

        uc1, uc2, uc3, uc4, uc5, uc6 = st.columns(6)
        uc1.metric("SKUs to Order",     len(uc_filtered))
        uc2.metric("Pre-Tax Total",     f"${uc_pretax:,.2f}")
        uc3.metric(f"Tax ({tax_rate*100:.3g}%)", f"${uc_tax:,.2f}")
        uc4.metric("Subtotal (Tax-In)", f"${uc_taxin:,.2f}")
        uc5.metric("Shipping",          f"${shipping_cost:,}" if shipping_cost > 0 else "—")
        uc6.metric("Grand Total",       f"${uc_grand:,.2f}")

        st.markdown("---")

        _uc_cat_costs  = uc_filtered.groupby('Classification')['Est Cost'].sum().sort_values(ascending=False)
        _uc_tier_costs = uc_filtered.groupby('Tier')['Est Cost'].sum()
        _CAT_COLORS  = {'Flower':'#89d4f5','Pre-Roll':'#b8e0ff','Vapes':'#7ec8e3','Edibles':'#f5c842','Concentrates':'#e07b39','Beverages':'#6dd6a0','Capsules':'#b89ef5','Oil':'#f5a0c8','Topicals':'#8fd48f','Seeds':'#d4c089'}
        _TIER_COLORS = {'A':'#4caf80','B':'#8bc34a','C':'#ffc107','D':'#ef5350'}

        _uc_c1, _uc_c2 = st.columns([3, 2])
        with _uc_c1:
            _uc_fig_cat = go.Figure(go.Bar(
                x=_uc_cat_costs.values, y=_uc_cat_costs.index, orientation='h',
                marker_color=[_CAT_COLORS.get(c,'#89d4f5') for c in _uc_cat_costs.index],
                text=[f"${v:,.0f}" for v in _uc_cat_costs.values], textposition='outside',
                hovertemplate='%{y}: $%{x:,.2f}<extra></extra>',
            ))
            _uc_fig_cat.update_layout(
                title=dict(text="Budget by Category", font=dict(size=13, color='#e8e8f0')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#e8e8f0', size=11),
                xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
                yaxis=dict(autorange='reversed', tickfont=dict(size=11)),
                margin=dict(l=10, r=60, t=36, b=10),
                height=max(200, len(_uc_cat_costs) * 32 + 60),
            )
            st.plotly_chart(_uc_fig_cat, use_container_width=True, config={'displayModeBar': False})
        with _uc_c2:
            _tier_label_map = {'A':'A — Fast','B':'B — Mid','C':'C — Slow','D':'D — Very Slow'}
            _uc_fig_tier = go.Figure(go.Pie(
                labels=[_tier_label_map.get(t, t) for t in _uc_tier_costs.index],
                values=_uc_tier_costs.values,
                marker_colors=[_TIER_COLORS.get(t,'#aaa') for t in _uc_tier_costs.index],
                hole=0.55, textinfo='label+percent',
                hovertemplate='%{label}: $%{value:,.2f}<extra></extra>',
                textfont=dict(size=11),
            ))
            _uc_fig_tier.update_layout(
                title=dict(text="Budget by Tier", font=dict(size=13, color='#e8e8f0')),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#e8e8f0'), showlegend=False,
                margin=dict(l=10, r=10, t=36, b=10),
                height=max(200, len(_uc_cat_costs) * 32 + 60),
            )
            st.plotly_chart(_uc_fig_tier, use_container_width=True, config={'displayModeBar': False})

        st.markdown("---")

        if uc_filtered.empty:
            st.info("No items match the selected tiers.")
        else:
            uc_display_cols = {
                'Classification':'Category','Product':'Product','Risk':'Risk','Net Days':'Net Days','Tier':'Tier',
                'Days Left':'Days Left','In Stock Qty':'In Stock','On Order':'On Order',
                'Sales (7 Days)':'7d Sales','Sales (30 Days)':'30d Sales',
                'Weekly Vel':'Wkly Vel','Cases':'Cases','Suggest':'Units',
                'Unit Price':'Unit Price','Est Cost':'Est Cost','Supplier Sku':'OCS Variant #'
            }
            def color_risk(val):
                return {'Critical':'background-color:#FFC7CE;color:#9C0006',
                        'At Risk': 'background-color:#FFEB9C;color:#9C6500',
                        'Stable':  'background-color:#C6EFCE;color:#276221'}.get(val,'')
            uc_show = uc_filtered[[c for c in uc_display_cols if c in uc_filtered.columns]].rename(columns=uc_display_cols).copy()
            uc_show['Unit Price'] = uc_show['Unit Price'].map(lambda x: f'${x:,.2f}' if pd.notna(x) else '—')
            uc_show['Est Cost']   = uc_show['Est Cost'].map(lambda x: f'${x:,.2f}' if pd.notna(x) else '—')
            uc_show['Days Left']  = uc_show['Days Left'].round(1)
            st.dataframe(
                uc_show.style.map(color_risk, subset=['Risk']).map(color_tier, subset=['Tier']).map(color_days, subset=['Days Left']),
                hide_index=True, use_container_width=True, height=500
            )

            st.markdown("---")
            uc_today = date.today()
            uc_upload = uc_filtered[['Supplier Sku','Cases']].copy()
            uc_upload.columns = ['SKU','Quantity']
            uc_buf = io.BytesIO()
            with pd.ExcelWriter(uc_buf, engine='openpyxl') as writer:
                uc_upload.to_excel(writer, sheet_name='MasterCatalogue', index=False)
            uc_buf.seek(0)
            st.download_button(
                label="📤 Download OCS Upload File (.xlsx)",
                data=uc_buf,
                file_name=f'OCS_SuggestedOrder_{uc_today.strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )

# ══════════════════════════════════════════════════════════════
# TAB 3: MENU BUILDER
# ══════════════════════════════════════════════════════════════
with tab3:
    st.markdown("### 🗂️ Menu Builder")
    st.caption("Set how many SKUs you want on your shelf per category, size, and strain. The tool shows your gaps and suggests what to order.")

    today2 = date.today()
    STRAINS   = ['Indica','Sativa','Hybrid','Blend']
    ALL_CATS  = ['Flower','Pre-Roll','Vapes','Edibles','Concentrates','Beverages','Capsules','Oil','Topicals','Seeds']
    CAT_EMOJIS = {
        'Flower':'🌸','Pre-Roll':'🚬','Vapes':'💨','Edibles':'🍬',
        'Concentrates':'🔬','Beverages':'🥤','Capsules':'💊',
        'Oil':'🫙','Topicals':'🧴','Seeds':'🌱',
    }

    def sort_size_key(s):
        import re as _re
        m = _re.match(r'^(\d+)x(\d+\.?\d*)g?$', str(s))
        if m: return (0, int(m.group(1)), float(m.group(2)))
        m = _re.match(r'^(\d+\.?\d*)g$', str(s))
        if m: return (1, 0, float(m.group(1)))
        m = _re.match(r'^(\d+\.?\d*)ml$', str(s))
        if m: return (2, 0, float(m.group(1)))
        m = _re.match(r'^(\d+\.?\d*)mg$', str(s))
        if m: return (3, 0, float(m.group(1)))
        return (99, 0, 0)

    # Detect sizes per category
    cat_sizes = {}
    for cat in ALL_CATS:
        raw = merged_raw[(merged_raw['Classification']==cat) & merged_raw['Product Size'].notna()]['Product Size'].unique().tolist()
        try:    cat_sizes[cat] = sorted(raw, key=sort_size_key)
        except: cat_sizes[cat] = sorted(raw)

    # Sub-categories per category (auto-detected from OCS Sub-Category column)
    has_subcat  = 'Sub-Category' in merged_raw.columns
    subcat_map  = {}          # cat -> [subcats]
    subcat_sizes = {}         # (cat, subcat) -> [sizes]
    if has_subcat:
        for _cat in ALL_CATS:
            _sc_vals = merged_raw[
                (merged_raw['Classification']==_cat) & (merged_raw['Sub-Category']!='')
            ]['Sub-Category'].dropna().unique().tolist()
            if _sc_vals:
                subcat_map[_cat] = sorted(_sc_vals)
                for _sc in _sc_vals:
                    _raw = merged_raw[
                        (merged_raw['Classification']==_cat) &
                        (merged_raw['Sub-Category']==_sc) &
                        merged_raw['Product Size'].notna()
                    ]['Product Size'].unique().tolist()
                    try:    subcat_sizes[(_cat, _sc)] = sorted(_raw, key=sort_size_key)
                    except: subcat_sizes[(_cat, _sc)] = sorted(_raw)

    # ── target settings ───────────────────────────────────────
    st.markdown("#### Target SKU Counts")
    mb_targets = {}        # (cat, size, strain) for cats without sub-categories
    mb_subcat_targets = {} # (cat, subcat, size, strain) for cats with sub-categories

    for cat in ALL_CATS:
        sizes  = cat_sizes[cat]
        emoji  = CAT_EMOJIS[cat]
        subcats = subcat_map.get(cat, [])

        if subcats:
            with st.expander(f"{emoji} {cat} — by sub-category, size & strain", expanded=False):
                st.caption(f"{len(subcats)} sub-categories detected: {', '.join(subcats)}")
                for sc in subcats:
                    sc_sizes = subcat_sizes.get((cat, sc), [])
                    st.markdown(f"**🔹 {sc}**")
                    if not sc_sizes:
                        st.caption(f"No sizes detected for {sc}.")
                        continue
                    hdr = st.columns([2]+[1]*4)
                    hdr[0].markdown("**Size**")
                    for i, s in enumerate(STRAINS): hdr[i+1].markdown(f"**{s}**")
                    for size in sc_sizes:
                        rc = st.columns([2]+[1]*4)
                        rc[0].markdown(f"`{size}`")
                        for i, strain in enumerate(STRAINS):
                            default = 0
                            if cat == 'Flower' and sc == 'Whole Flower':
                                default = 4 if size in ['3.5g','7g'] and strain in ['Indica','Sativa'] else \
                                          2 if size in ['3.5g','7g'] and strain in ['Hybrid','Blend'] else 1
                            mb_subcat_targets[(cat, sc, size, strain)] = rc[i+1].number_input(
                                "", value=default, min_value=0, max_value=50,
                                key=f"mb_{cat}_{sc}_{size}_{strain}",
                                label_visibility="collapsed")
        else:
            with st.expander(f"{emoji} {cat} — by size & strain", expanded=(cat=='Flower')):
                if not sizes:
                    st.caption(f"No {cat} sizes detected in your catalogue.")
                    continue
                st.caption(f"{len(sizes)} sizes detected from your catalogue.")
                hdr = st.columns([2]+[1]*4)
                hdr[0].markdown("**Size**")
                for i, s in enumerate(STRAINS): hdr[i+1].markdown(f"**{s}**")
                for size in sizes:
                    rc = st.columns([2]+[1]*4)
                    rc[0].markdown(f"`{size}`")
                    for i, strain in enumerate(STRAINS):
                        default = 0
                        if cat == 'Flower':
                            default = 4 if size in ['3.5g','7g'] and strain in ['Indica','Sativa'] else \
                                      2 if size in ['3.5g','7g'] and strain in ['Hybrid','Blend'] else 1
                        mb_targets[(cat, size, strain)] = rc[i+1].number_input(
                            "", value=default, min_value=0, max_value=50,
                            key=f"mb_{cat}_{size}_{strain}", label_visibility="collapsed")

    st.markdown("---")

    # ── shelf counts ──────────────────────────────────────────
    on_shelf     = merged_raw[merged_raw['In Stock Qty'] > 0].copy()
    shelf_counts = on_shelf.groupby(['Classification','Product Size','Strain'])['SKU'].nunique().reset_index()
    shelf_counts.columns = ['Category','Size','Strain','On Shelf']

    # Sub-category shelf counts (for all cats that have sub-categories)
    subcat_shelf = pd.DataFrame()
    if has_subcat and subcat_map:
        sc_cats = list(subcat_map.keys())
        sc_raw = on_shelf[on_shelf['Classification'].isin(sc_cats)]
        if not sc_raw.empty:
            subcat_shelf = sc_raw.groupby(['Classification','Sub-Category','Product Size','Strain'])['SKU'].nunique().reset_index()
            subcat_shelf.columns = ['Category','Sub-Category','Size','Strain','On Shelf']

    # ── cart boost: each cart SKU counts as +1 on-shelf for its (cat, size, strain) ──
    if 'menu_cart' not in st.session_state:
        st.session_state['menu_cart'] = {}
    _cart_boost       = {}   # (cat, size, strain) -> int
    _cart_subcat_boost = {}  # (cat, subcat, size, strain) -> int
    for _cb_item in st.session_state['menu_cart'].values():
        _cb_key = (_cb_item.get('Category',''), _cb_item.get('Size',''), _cb_item.get('Strain',''))
        _cart_boost[_cb_key] = _cart_boost.get(_cb_key, 0) + 1
        _cb_sc = _cb_item.get('Sub-Type','')
        if _cb_sc:
            _cb_skey = (_cb_item.get('Category',''), _cb_sc, _cb_item.get('Size',''), _cb_item.get('Strain',''))
            _cart_subcat_boost[_cb_skey] = _cart_subcat_boost.get(_cb_skey, 0) + 1

    # ── gap helpers ───────────────────────────────────────────
    def highlight_gap(val):
        if val > 0: return 'background-color:#FFC7CE;font-weight:bold'
        return 'background-color:#C6EFCE'

    def show_size_strain_grid(cat, sizes):
        rows = []
        for size in sizes:
            for strain in STRAINS:
                tgt = mb_targets.get((cat, size, strain), 0)
                cur = int(shelf_counts[
                    (shelf_counts['Category']==cat) &
                    (shelf_counts['Size']==size) &
                    (shelf_counts['Strain']==strain)
                ]['On Shelf'].sum()) + _cart_boost.get((cat, size, strain), 0)
                rows.append({'Size': size, 'Strain': strain, 'Target': tgt, 'On Shelf': cur, 'Gap': max(0, tgt-cur)})
        if not rows:
            return pd.DataFrame(0, index=sizes, columns=STRAINS)
        df = pd.DataFrame(rows)
        ps = df.pivot_table(index='Size', columns='Strain', values='On Shelf', fill_value=0).reindex(sizes).fillna(0).astype(int)
        pt = df.pivot_table(index='Size', columns='Strain', values='Target',   fill_value=0).reindex(sizes).fillna(0).astype(int)
        pg = df.pivot_table(index='Size', columns='Strain', values='Gap',      fill_value=0).reindex(sizes).fillna(0).astype(int)
        for piv in [ps, pt, pg]:
            for col in STRAINS:
                if col not in piv.columns: piv[col] = 0
        gc1, gc2, gc3 = st.columns(3)
        gc1.caption("On Shelf now");        gc1.dataframe(ps[STRAINS], use_container_width=True)
        gc2.caption("Your target");         gc2.dataframe(pt[STRAINS], use_container_width=True)
        gc3.caption("Gap (need to order)"); gc3.dataframe(pg[STRAINS].style.map(highlight_gap), use_container_width=True)
        return pg[STRAINS]

    def show_subcat_grid(cat, sc, sc_sizes):
        rows = []
        for size in sc_sizes:
            for strain in STRAINS:
                tgt = mb_subcat_targets.get((cat, sc, size, strain), 0)
                if subcat_shelf.empty:
                    cur = 0
                else:
                    cur = int(subcat_shelf[
                        (subcat_shelf['Category']==cat) &
                        (subcat_shelf['Sub-Category']==sc) &
                        (subcat_shelf['Size']==size) &
                        (subcat_shelf['Strain']==strain)
                    ]['On Shelf'].sum())
                cur += _cart_subcat_boost.get((cat, sc, size, strain), 0)
                rows.append({'Size': size, 'Strain': strain, 'Target': tgt, 'On Shelf': cur, 'Gap': max(0, tgt-cur)})
        if not rows:
            return pd.DataFrame(0, index=sc_sizes, columns=STRAINS)
        df = pd.DataFrame(rows)
        ps = df.pivot_table(index='Size', columns='Strain', values='On Shelf', fill_value=0).reindex(sc_sizes).fillna(0).astype(int)
        pt = df.pivot_table(index='Size', columns='Strain', values='Target',   fill_value=0).reindex(sc_sizes).fillna(0).astype(int)
        pg = df.pivot_table(index='Size', columns='Strain', values='Gap',      fill_value=0).reindex(sc_sizes).fillna(0).astype(int)
        for piv in [ps, pt, pg]:
            for col in STRAINS:
                if col not in piv.columns: piv[col] = 0
        gc1, gc2, gc3 = st.columns(3)
        gc1.caption("On Shelf now");        gc1.dataframe(ps[STRAINS], use_container_width=True)
        gc2.caption("Your target");         gc2.dataframe(pt[STRAINS], use_container_width=True)
        gc3.caption("Gap (need to order)"); gc3.dataframe(pg[STRAINS].style.map(highlight_gap), use_container_width=True)
        return pg[STRAINS]

    # ── gap display ───────────────────────────────────────────
    st.markdown("#### Current Shelf vs Target")
    cat_pivot_gaps = {}        # cat -> gap pivot  (no sub-categories)
    subcat_gaps    = {}        # (cat, subcat) -> gap pivot

    for cat in ALL_CATS:
        subcats = subcat_map.get(cat, [])
        if subcats:
            any_tgt = any(mb_subcat_targets.get((cat, sc, s, st_), 0) > 0
                          for sc in subcats
                          for s in subcat_sizes.get((cat, sc), [])
                          for st_ in STRAINS)
            if not any_tgt:
                continue
            st.markdown(f"**{CAT_EMOJIS[cat]} {cat}**")
            for sc in subcats:
                sc_sizes = subcat_sizes.get((cat, sc), [])
                if not sc_sizes: continue
                if not any(mb_subcat_targets.get((cat, sc, s, st_), 0) > 0 for s in sc_sizes for st_ in STRAINS):
                    continue
                st.markdown(f"*🔹 {sc}*")
                subcat_gaps[(cat, sc)] = show_subcat_grid(cat, sc, sc_sizes)
        else:
            sizes = cat_sizes[cat]
            if not sizes: continue
            if not any(mb_targets.get((cat, s, st_), 0) > 0 for s in sizes for st_ in STRAINS): continue
            st.markdown(f"**{CAT_EMOJIS[cat]} {cat}**")
            cat_pivot_gaps[cat] = show_size_strain_grid(cat, sizes)

    # ── menu export ───────────────────────────────────────────
    def build_menu_workbook(_suggestions):
        wb = Workbook()
        PINK_FILL2  = _p('FFC7CE')
        GREEN_FILL2 = _p('C6EFCE')

        # ── Sheet 1: Shelf vs Target ──────────────────────────
        ws1 = wb.active
        ws1.title = "Shelf vs Target"
        _ncols = 1 + len(STRAINS) * 3
        _col_letter = ws1.cell(row=1, column=_ncols).column_letter

        # Title row
        _tc = ws1.cell(row=1, column=1, value="Menu Builder — Shelf vs Target")
        _tc.font = Font(color='FFFFFF', bold=True, size=13)
        _tc.fill = _p('0F2A3D')
        _tc.alignment = _a('center')
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_ncols)
        _dc = ws1.cell(row=2, column=1, value=f"Generated: {today2.strftime('%B %d, %Y')}")
        _dc.font = Font(color='808080', size=9, italic=True)
        _dc.alignment = _a('right')
        ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_ncols)
        ws1.column_dimensions['A'].width = 16
        for _ci in range(2, _ncols + 1):
            ws1.column_dimensions[ws1.cell(row=1, column=_ci).column_letter].width = 12

        def _write_grid(ws, start_row, label, sizes, shelf_data, target_data, is_subcat=False):
            r = start_row
            # Section header
            _hc = ws.cell(row=r, column=1, value=label)
            _hc.font = Font(color='FFFFFF', bold=True, size=11)
            _hc.fill = _p('2E75B6') if is_subcat else _p('1F4E79')
            _hc.alignment = _a('center')
            _hc.border = BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_ncols)
            r += 1
            # Group headers
            ws.cell(row=r, column=1, value='Size').font = Font(bold=True, size=9)
            ws.cell(row=r, column=1).fill = _p('D9D9D9')
            ws.cell(row=r, column=1).alignment = _a('center')
            ws.cell(row=r, column=1).border = BORDER
            for _gi, (_gc, _glabel, _gfill) in enumerate([
                (2, 'On Shelf Now', _p('2167A0')),
                (2+len(STRAINS), 'Target SKUs', _p('375623')),
                (2+len(STRAINS)*2, 'Gap (to order)', _p('843C0C')),
            ]):
                _ghc = ws.cell(row=r, column=_gc, value=_glabel)
                _ghc.font = Font(color='FFFFFF', bold=True, size=9)
                _ghc.fill = _gfill
                _ghc.alignment = _a('center')
                _ghc.border = BORDER
                ws.merge_cells(start_row=r, start_column=_gc, end_row=r, end_column=_gc+len(STRAINS)-1)
            r += 1
            # Strain sub-headers
            ws.cell(row=r, column=1, value='').border = BORDER
            ws.cell(row=r, column=1).fill = _p('D9D9D9')
            for _gs in [2, 2+len(STRAINS), 2+len(STRAINS)*2]:
                for _si, _sn in enumerate(STRAINS):
                    _shc = ws.cell(row=r, column=_gs+_si, value=_sn)
                    _shc.font = Font(bold=True, size=9)
                    _shc.alignment = _a('center')
                    _shc.border = BORDER
                    _shc.fill = _p('F2F2F2')
            r += 1
            # Data rows
            for size in sizes:
                ws.cell(row=r, column=1, value=size).font = Font(bold=True, size=10)
                ws.cell(row=r, column=1).alignment = _a('center')
                ws.cell(row=r, column=1).border = BORDER
                ws.cell(row=r, column=1).fill = _p('EBEBEB')
                for _si, _sn in enumerate(STRAINS):
                    sv = int(shelf_data.get((size, _sn), 0))
                    tv = int(target_data.get((size, _sn), 0))
                    gv = max(0, tv - sv)
                    _sc2 = ws.cell(row=r, column=2+_si, value=sv)
                    _sc2.alignment = _a('center'); _sc2.border = BORDER; _sc2.fill = _p('EBF3FB')
                    _tc2 = ws.cell(row=r, column=2+len(STRAINS)+_si, value=tv)
                    _tc2.alignment = _a('center'); _tc2.border = BORDER
                    _tc2.fill = _p('EBF7E6') if tv > 0 else _p('F2F2F2')
                    _gc2 = ws.cell(row=r, column=2+len(STRAINS)*2+_si, value=gv)
                    _gc2.alignment = _a('center'); _gc2.border = BORDER
                    _gc2.fill = PINK_FILL2 if gv > 0 else GREEN_FILL2
                    if gv > 0: _gc2.font = Font(bold=True, color='9C0006', size=10)
                r += 1
            # Totals row
            ws.cell(row=r, column=1, value='TOTAL GAP').fill = HDR_FILL
            ws.cell(row=r, column=1).font = HDR_FONT
            ws.cell(row=r, column=1).alignment = _a('center')
            ws.cell(row=r, column=1).border = BORDER
            for _si, _sn in enumerate(STRAINS):
                ws.cell(row=r, column=2+_si, value='').border = BORDER; ws.cell(row=r, column=2+_si).fill = HDR_FILL
                ws.cell(row=r, column=2+len(STRAINS)+_si, value='').border = BORDER; ws.cell(row=r, column=2+len(STRAINS)+_si).fill = HDR_FILL
                _col_g = sum(max(0, int(target_data.get((s, _sn), 0)) - int(shelf_data.get((s, _sn), 0))) for s in sizes)
                _gc3 = ws.cell(row=r, column=2+len(STRAINS)*2+_si, value=_col_g)
                _gc3.alignment = _a('center'); _gc3.border = BORDER
                _gc3.fill = PINK_FILL2 if _col_g > 0 else GREEN_FILL2
                if _col_g > 0: _gc3.font = Font(bold=True, color='9C0006', size=10)
            return r + 2

        ws1_row = 3
        for _cat in ALL_CATS:
            _emoji = CAT_EMOJIS.get(_cat, '')
            _subcats = subcat_map.get(_cat, [])
            if _subcats:
                if not any(mb_subcat_targets.get((_cat, _sc, _s, _st), 0) > 0
                           for _sc in _subcats for _s in subcat_sizes.get((_cat, _sc), []) for _st in STRAINS):
                    continue
                _ch = ws1.cell(row=ws1_row, column=1, value=f"{_emoji} {_cat}")
                _ch.font = Font(color='FFFFFF', bold=True, size=12)
                _ch.fill = _p('0D3B66')
                _ch.alignment = _a('center')
                _ch.border = BORDER
                ws1.merge_cells(start_row=ws1_row, start_column=1, end_row=ws1_row, end_column=_ncols)
                ws1_row += 1
                for _sc in _subcats:
                    _sc_sizes = subcat_sizes.get((_cat, _sc), [])
                    if not _sc_sizes: continue
                    if not any(mb_subcat_targets.get((_cat, _sc, _s, _st), 0) > 0 for _s in _sc_sizes for _st in STRAINS):
                        continue
                    _sd = {}
                    for _s in _sc_sizes:
                        for _st in STRAINS:
                            _sv = int(subcat_shelf[
                                (subcat_shelf['Category']==_cat) & (subcat_shelf['Sub-Category']==_sc) &
                                (subcat_shelf['Size']==_s) & (subcat_shelf['Strain']==_st)
                            ]['On Shelf'].sum()) if not subcat_shelf.empty else 0
                            _sd[(_s, _st)] = _sv + _cart_subcat_boost.get((_cat, _sc, _s, _st), 0)
                    _td = {(_s, _st): mb_subcat_targets.get((_cat, _sc, _s, _st), 0) for _s in _sc_sizes for _st in STRAINS}
                    ws1_row = _write_grid(ws1, ws1_row, f"🔹 {_sc}", _sc_sizes, _sd, _td, is_subcat=True)
            else:
                _sizes = cat_sizes[_cat]
                if not _sizes: continue
                if not any(mb_targets.get((_cat, _s, _st), 0) > 0 for _s in _sizes for _st in STRAINS): continue
                _sd = {}
                for _s in _sizes:
                    for _st in STRAINS:
                        _sv = int(shelf_counts[
                            (shelf_counts['Category']==_cat) & (shelf_counts['Size']==_s) & (shelf_counts['Strain']==_st)
                        ]['On Shelf'].sum()) + _cart_boost.get((_cat, _s, _st), 0)
                        _sd[(_s, _st)] = _sv
                _td = {(_s, _st): mb_targets.get((_cat, _s, _st), 0) for _s in _sizes for _st in STRAINS}
                ws1_row = _write_grid(ws1, ws1_row, f"{_emoji} {_cat}", _sizes, _sd, _td)

        # ── Sheet 2: Gap Suggestions ──────────────────────────
        ws2 = wb.create_sheet("Gap Suggestions")
        _sug_hdrs = ['Product','Category','Sub-Type','Size','Strain','OCS Variant #','OCS Item #',
                     'Prev. Sold (60d)','Cases','Unit Price','Est. Cost','Status']
        _sug_ws   = [45, 14, 14, 10, 10, 18, 14, 18, 8, 12, 12, 22]
        _t2 = ws2.cell(row=1, column=1, value="Gap Suggestions — SKUs to Order")
        _t2.font = Font(color='FFFFFF', bold=True, size=13)
        _t2.fill = _p('0F2A3D')
        _t2.alignment = _a('center')
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_sug_hdrs))
        _d2 = ws2.cell(row=2, column=1, value=f"Generated: {today2.strftime('%B %d, %Y')}")
        _d2.font = Font(color='808080', size=9, italic=True)
        _d2.alignment = _a('right')
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_sug_hdrs))
        for _ci, (_hdr, _w) in enumerate(zip(_sug_hdrs, _sug_ws), start=1):
            _hc2 = ws2.cell(row=4, column=_ci, value=_hdr)
            _hc2.font = HDR_FONT; _hc2.fill = HDR_FILL
            _hc2.alignment = _a('center'); _hc2.border = BORDER
            ws2.column_dimensions[_hc2.column_letter].width = _w
        ws2.freeze_panes = 'A5'
        _sug_row = 5
        for _sg in _suggestions:
            _up  = _sg.get('Unit Price') or 0
            _est = _sg['Cases'] * _sg['Pack Size'] * _up if _up else 0
            _status = 'NEW — never ordered' if _sg.get('Is New') else 'Previously carried'
            _vals = [_sg['Product'], _sg['Category'], _sg.get('Sub-Type',''), _sg['Size'], _sg['Strain'],
                     _sg['OCS Variant #'], _sg['OCS Item #'], _sg.get('Prev. Sold (60d)', 0),
                     _sg['Cases'], _up if _up else None, _est if _est else None, _status]
            _rfill = _p('FFF8E1') if _sg.get('Is New') else _p('F0F8FF')
            for _ci, _val in enumerate(_vals, start=1):
                _rc2 = ws2.cell(row=_sug_row, column=_ci, value=_val)
                _rc2.fill = _rfill; _rc2.border = BORDER
                _rc2.alignment = _a('center' if _ci >= 8 else 'left')
                if _ci == 10 and _val: _rc2.number_format = '$#,##0.00'
                if _ci == 11 and _val: _rc2.number_format = '$#,##0.00'; _rc2.font = Font(bold=True, size=10)
            _sug_row += 1
        if _suggestions:
            _tot_est = sum((_sg['Cases'] * _sg['Pack Size'] * (_sg.get('Unit Price') or 0)) for _sg in _suggestions)
            _tr = ws2.cell(row=_sug_row, column=1, value='TOTAL')
            _tr.font = HDR_FONT; _tr.fill = HDR_FILL; _tr.border = BORDER
            for _ci in range(2, len(_sug_hdrs)+1):
                ws2.cell(row=_sug_row, column=_ci).fill = HDR_FILL
                ws2.cell(row=_sug_row, column=_ci).border = BORDER
            _tot_c = ws2.cell(row=_sug_row, column=11, value=_tot_est)
            _tot_c.font = Font(color='FFFFFF', bold=True, size=10)
            _tot_c.fill = HDR_FILL; _tot_c.alignment = _a('center')
            _tot_c.number_format = '$#,##0.00'; _tot_c.border = BORDER
        elif not _suggestions:
            ws2.cell(row=5, column=1, value='No gaps found — shelf is fully stocked to target.').font = Font(italic=True, color='808080', size=10)

        _buf = io.BytesIO()
        wb.save(_buf)
        _buf.seek(0)
        return _buf

    # ── suggestions ───────────────────────────────────────────
    total_gaps = (sum(int(pg.values.sum()) for pg in cat_pivot_gaps.values()) +
                  sum(int(pg.values.sum()) for pg in subcat_gaps.values()))
    suggestions = []
    if not cat_pivot_gaps and not subcat_gaps or total_gaps == 0:
        st.success("✅ Your shelf is fully stocked to target across all categories!")
        if cat_pivot_gaps or subcat_gaps:
            _mb_buf = build_menu_workbook([])
            st.download_button(
                "📥 Download Menu Report (.xlsx)",
                data=_mb_buf,
                file_name=f'MenuReport_{today2.strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
    else:
        st.markdown("---")
        st.markdown("#### 💡 Suggested SKUs to Fill Gaps")
        st.caption("SKUs you've sold before (priority) or available on OCS — order 1 case each to fill gaps.")

        suggestions = []
        _on_shelf_skus = set(merged_raw[merged_raw['In Stock Qty'] > 0]['Supplier Sku'].tolist())
        _ocs_has_class = 'Classification' in ocs_df.columns

        def _append_suggestion(cat, sc, size, strain, r, prev=0, is_new=False, sku_col='Supplier Sku'):
            suggestions.append({
                'Category': cat, 'Sub-Type': sc, 'Size': size, 'Strain': strain,
                'Product': r.get('Product',''),
                'OCS Variant #': r.get(sku_col, r.get('OCS Variant Number','')),
                'OCS Item #': str(r['OCS Item Number']).split('.')[0] if pd.notna(r.get('OCS Item Number')) else '—',
                'Cases': 1, 'Pack Size': int(r.get('Pack Size', 1)),
                'Unit Price': r.get('Unit Price') if pd.notna(r.get('Unit Price')) else None,
                'Prev. Sold (60d)': prev,
                'Is New': is_new,
            })

        def _ocs_supplement(cat, sc, size, strain, needed, exclude_skus):
            """Pull brand-new OCS SKUs to fill any remaining gap slots."""
            if needed <= 0 or not _ocs_has_class: return
            _mask = (
                (ocs_df['Classification']==cat) &
                (ocs_df['Product Size']==size) &
                (ocs_df['Strain']==strain) &
                (ocs_df['Stock Status']=='YES') &
                ~ocs_df['OCS Variant Number'].isin(exclude_skus) &
                ~ocs_df['OCS Variant Number'].isin(_on_shelf_skus)
            )
            if sc and 'Sub-Category' in ocs_df.columns:
                _mask &= (ocs_df['Sub-Category']==sc)
            for _, r in ocs_df[_mask].head(needed).iterrows():
                _append_suggestion(cat, sc, size, strain, r, prev=0, is_new=True,
                                   sku_col='OCS Variant Number')
                exclude_skus.add(r['OCS Variant Number'])

        # Flat-category suggestions
        for cat, pg in cat_pivot_gaps.items():
            for size in cat_sizes[cat]:
                if size not in pg.index: continue
                for strain in STRAINS:
                    if strain not in pg.columns: continue
                    gap = int(pg.at[size, strain])
                    if gap <= 0: continue
                    candidates = merged_raw[
                        (merged_raw['Classification']==cat) &
                        (merged_raw['Product Size']==size) &
                        (merged_raw['Strain']==strain) &
                        (merged_raw['In Stock Qty']==0) &
                        (merged_raw['Stock Status']=='YES')
                    ].copy()
                    candidates['_sold'] = candidates['Sales (60 Days)'] > 0
                    candidates = candidates.sort_values(['_sold','Sales (60 Days)'], ascending=[False,False])
                    _used = set()
                    for _, r in candidates.head(gap).iterrows():
                        _prev = int(r['Sales (60 Days)']) if pd.notna(r.get('Sales (60 Days)')) else 0
                        _append_suggestion(cat, '', size, strain, r, prev=_prev, is_new=(_prev==0))
                        _used.add(r['Supplier Sku'])
                    _remaining = gap - len(_used)
                    _ocs_supplement(cat, '', size, strain, _remaining,
                                    _used | {s['OCS Variant #'] for s in suggestions})

        # Sub-category suggestions (Flower, Vapes, Topicals, etc.)
        for (cat, sc), pg in subcat_gaps.items():
            sc_sizes = subcat_sizes.get((cat, sc), [])
            for size in sc_sizes:
                if size not in pg.index: continue
                for strain in STRAINS:
                    if strain not in pg.columns: continue
                    gap = int(pg.at[size, strain])
                    if gap <= 0: continue
                    cand_mask = (
                        (merged_raw['Classification']==cat) &
                        (merged_raw['Product Size']==size) &
                        (merged_raw['Strain']==strain) &
                        (merged_raw['In Stock Qty']==0) &
                        (merged_raw['Stock Status']=='YES')
                    )
                    if has_subcat:
                        cand_mask &= (merged_raw['Sub-Category']==sc)
                    candidates = merged_raw[cand_mask].copy()
                    candidates['_sold'] = candidates['Sales (60 Days)'] > 0
                    candidates = candidates.sort_values(['_sold','Sales (60 Days)'], ascending=[False,False])
                    _used = set()
                    for _, r in candidates.head(gap).iterrows():
                        _prev = int(r['Sales (60 Days)']) if pd.notna(r.get('Sales (60 Days)')) else 0
                        _append_suggestion(cat, sc, size, strain, r, prev=_prev, is_new=(_prev==0))
                        _used.add(r['Supplier Sku'])
                    _remaining = gap - len(_used)
                    _ocs_supplement(cat, sc, size, strain, _remaining,
                                    _used | {s['OCS Variant #'] for s in suggestions})

        if suggestions:
            _has_subtype = any(s.get('Sub-Type') for s in suggestions)
            _has_new     = any(s.get('Is New') for s in suggestions)

            # ── legend ─────────────────────────────────────────
            _leg_parts = ["<small>"]
            _leg_parts.append("<span style='background:rgba(137,212,245,0.15);border:1px solid rgba(137,212,245,0.4);color:#89d4f5;padding:1px 7px;border-radius:4px;font-size:11px'>Previously carried</span>")
            if _has_new:
                _leg_parts.append("&nbsp;&nbsp;<span style='background:rgba(255,200,60,0.18);border:1px solid rgba(255,200,60,0.55);color:#ffc83c;padding:1px 7px;border-radius:4px;font-size:11px'>NEW — never ordered before</span>")
            _leg_parts.append("</small>")
            st.markdown(" ".join(_leg_parts), unsafe_allow_html=True)
            st.markdown("<div style='margin-bottom:6px'></div>", unsafe_allow_html=True)

            # ── header ─────────────────────────────────────────
            _hw = [3.2, 0.85, 0.65, 0.6, 0.55, 0.75, 0.5]
            _hh = st.columns(_hw)
            for _lbl, _col in zip(["Product","Category · Size","Strain","Prev Sold","Cases","Est. Cost",""], _hh):
                _col.markdown(f"<small><b>{_lbl}</b></small>", unsafe_allow_html=True)
            st.markdown("<hr style='margin:2px 0 6px 0;border-color:rgba(137,212,245,0.15)'>",
                        unsafe_allow_html=True)

            # ── suggestion rows ─────────────────────────────────
            for _si, _sr in enumerate(suggestions):
                _sku     = _sr['OCS Variant #']
                _in_cart = _sku in st.session_state['menu_cart']
                _up      = _sr.get('Unit Price') or 0
                _is_new  = _sr.get('Is New', False)
                _rc      = st.columns(_hw)
                _label   = _sr['Product']
                if _has_subtype and _sr.get('Sub-Type'):
                    _label = f"[{_sr['Sub-Type']}] {_label}"
                # colour-coded product name badge
                if _is_new:
                    _badge_style = "background:rgba(255,200,60,0.18);border:1px solid rgba(255,200,60,0.55);color:#ffc83c;padding:2px 6px;border-radius:4px;font-size:11px"
                    _badge_tag   = "<span style='background:rgba(255,200,60,0.35);color:#fff8e0;padding:0px 5px;border-radius:3px;font-size:10px;margin-right:4px;font-weight:600'>NEW</span>"
                else:
                    _badge_style = "background:rgba(137,212,245,0.1);border:1px solid rgba(137,212,245,0.25);color:#c8eaf8;padding:2px 6px;border-radius:4px;font-size:11px"
                    _badge_tag   = ""
                _rc[0].markdown(f"<span style='{_badge_style}'>{_badge_tag}{_label[:45]}</span>", unsafe_allow_html=True)
                _rc[1].caption(f"{_sr['Category']} · {_sr['Size']}")
                _rc[2].caption(_sr['Strain'])
                _rc[3].caption(str(_sr.get('Prev. Sold (60d)', 0)))
                _qty = _rc[4].number_input("", value=st.session_state.get(f"sug_q_{_si}", 1),
                                           min_value=1, max_value=99,
                                           key=f"sug_q_{_si}", label_visibility="collapsed")
                _est = _qty * _sr['Pack Size'] * _up
                _rc[5].caption(f"${_est:,.2f}" if _est else "—")
                _btn_lbl = "✓" if _in_cart else "+"
                if _rc[6].button(_btn_lbl, key=f"sg_{_si}", use_container_width=True,
                                 help="Remove from cart" if _in_cart else "Add to cart"):
                    if _in_cart:
                        del st.session_state['menu_cart'][_sku]
                    else:
                        _entry = dict(_sr)
                        _entry['Cases'] = _qty
                        st.session_state['menu_cart'][_sku] = _entry
                    st.rerun()
        else:
            st.info("No matching SKUs found in the OCS catalogue to fill the gaps. Try adjusting your targets or check the catalogue for availability.")

        # ── menu report download ───────────────────────────────
        st.markdown("---")
        _mb_buf = build_menu_workbook(suggestions)
        st.download_button(
            "📥 Download Menu Report (.xlsx)",
            data=_mb_buf,
            file_name=f'MenuReport_{today2.strftime("%Y%m%d")}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            help="Download a formatted Excel report with your Shelf vs Target grids and Gap Suggestions list.",
        )

        # ── menu cart ──────────────────────────────────────────
        _cart = st.session_state.get('menu_cart', {})
        if _cart:
            st.markdown("---")
            st.markdown("#### 🛒 Added to Order")
            _cart_total = 0.0
            _ch = st.columns([3.2, 1.0, 0.7, 0.55, 0.75, 0.45])
            for _lbl, _c in zip(["Product","Category · Size","Strain","Cases","Est. Cost",""], _ch):
                _c.markdown(f"<small><b>{_lbl}</b></small>", unsafe_allow_html=True)
            st.markdown("<hr style='margin:2px 0 6px 0;border-color:rgba(137,212,245,0.15)'>",
                        unsafe_allow_html=True)
            for _csku, _ci in list(_cart.items()):
                _cup    = _ci.get('Unit Price') or 0
                _ccases = _ci.get('Cases', 1)
                _cest   = _ccases * _ci['Pack Size'] * _cup
                _cart_total += _cest
                _cr = st.columns([3.2, 1.0, 0.7, 0.55, 0.75, 0.45])
                _cr[0].caption(_ci['Product'][:50])
                _cr[1].caption(f"{_ci['Category']} · {_ci['Size']}")
                _cr[2].caption(_ci['Strain'])
                _new_c = _cr[3].number_input("", value=_ccases, min_value=1, max_value=99,
                                              key=f"cart_c_{_csku}", label_visibility="collapsed")
                if _new_c != _ccases:
                    st.session_state['menu_cart'][_csku]['Cases'] = _new_c
                    st.rerun()
                _cr[4].caption(f"${_cest:,.2f}" if _cest else "—")
                if _cr[5].button("✕", key=f"crm_{_csku}", use_container_width=True):
                    del st.session_state['menu_cart'][_csku]
                    st.rerun()
            st.caption(f"Menu additions: {len(_cart)} SKU(s) · ${_cart_total:,.2f}")

        # ── full combined order preview ─────────────────────────
        st.markdown("---")
        st.markdown("#### 📋 Full Order Preview")
        st.caption("Everything that will be on the OCS upload file — your replenishment order + any menu additions.")

        _full_rows = []
        if not order_df.empty:
            for _, _or in order_df.iterrows():
                _full_rows.append({
                    'Source': 'Auto',
                    'Product': _or.get('Product',''),
                    'Category': _or.get('Classification',''),
                    'Size': _or.get('Product Size',''),
                    'Strain': _or.get('Strain',''),
                    'Cases': int(_or.get('Cases', 1)),
                    'Unit Price': _or.get('Unit Price', None),
                    'Pack Size': int(_or.get('Pack Size', 1)),
                    'SKU': _or.get('Supplier Sku',''),
                })
        for _msku, _mi in _cart.items():
            _full_rows.append({
                'Source': 'Menu',
                'Product': _mi.get('Product',''),
                'Category': _mi.get('Category',''),
                'Size': _mi.get('Size',''),
                'Strain': _mi.get('Strain',''),
                'Cases': int(_mi.get('Cases',1)),
                'Unit Price': _mi.get('Unit Price', None),
                'Pack Size': int(_mi.get('Pack Size',1)),
                'SKU': _msku,
            })

        if _full_rows:
            _full_df = pd.DataFrame(_full_rows)
            _full_df['Est. Cost'] = (_full_df['Cases'] * _full_df['Pack Size'] *
                                      _full_df['Unit Price'].fillna(0)).round(2)
            _disp = _full_df[['Source','Product','Category','Size','Strain','Cases','Est. Cost']].copy()
            _disp['Est. Cost'] = _disp['Est. Cost'].map(lambda x: f'${x:,.2f}' if x > 0 else '—')
            st.dataframe(_disp, hide_index=True, use_container_width=True)

            _total_cost = _full_df['Est. Cost'].sum()
            st.markdown(f"**Total: ${_total_cost:,.2f}** &nbsp;·&nbsp; {len(_full_df)} SKU(s) &nbsp;·&nbsp; "
                        f"{int(_full_df['Cases'].sum())} cases",
                        unsafe_allow_html=True)

            _combined_upload = _full_df[['SKU','Cases']].rename(columns={'Cases':'Quantity'})
            _combined_buf = io.BytesIO()
            with pd.ExcelWriter(_combined_buf, engine='openpyxl') as _fw:
                _combined_upload.to_excel(_fw, sheet_name='MasterCatalogue', index=False)
            _combined_buf.seek(0)
            st.download_button(
                "📤 Download Full OCS Upload File (.xlsx)",
                data=_combined_buf,
                file_name=f'OCS_FullOrder_{today2.strftime("%Y%m%d")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        else:
            st.info("No order items yet — upload your files and set targets to generate an order.")

# ══════════════════════════════════════════════════════════════════════════
# TAB 4 — AI MENU OPTIMIZER
# ══════════════════════════════════════════════════════════════════════════
with tab4:
    import plotly.graph_objects as _go4

    _AM_CATS = ['Flower','Pre-Roll','Vapes','Edibles','Concentrates','Beverages','Capsules','Oil','Topicals','Seeds']
    _today4  = date.today()

    st.markdown("### 🧠 AI Menu Optimizer")
    st.caption(
        "Builds your ideal carried menu automatically — picks the highest-velocity SKUs "
        "while ensuring balanced coverage across categories, sizes, and strains."
    )

    if all_active.empty:
        st.info("Upload your Cova and OCS files to generate an AI-optimized menu.")
    else:
        # ── controls ──────────────────────────────────────────────────────
        _am_col1, _am_col2, _am_col3, _am_col4 = st.columns([1, 1.2, 1, 2])
        with _am_col1:
            _am_n = st.number_input(
                "Target SKUs", min_value=10, max_value=2000, value=150, step=10,
                help="How many unique SKUs you want to carry on your menu."
            )
        with _am_col2:
            _am_mode = st.selectbox(
                "Selection mode",
                ["Velocity + Coverage", "Pure Velocity", "Revenue-weighted"],
                help=(
                    "**Velocity + Coverage** — fills proportional minimums per category first, "
                    "then tops up with highest-velocity SKUs across all categories.\n\n"
                    "**Pure Velocity** — strictly the top-N fastest movers, no category floor.\n\n"
                    "**Revenue-weighted** — ranks by daily velocity × unit price (highest revenue per day first)."
                )
            )
        with _am_col3:
            _am_tiers = st.multiselect(
                "Tiers to include", ['A','B','C','D'], default=['A','B','C'],
                help="Tier A = fastest movers, D = very slow. Typically exclude D for your core menu."
            )
        with _am_col4:
            _am_cats = st.multiselect(
                "Categories", _AM_CATS, default=_AM_CATS,
                label_visibility="visible"
            )

        # ── filter pool ───────────────────────────────────────────────────
        _am_pool = all_active[
            all_active['Classification'].isin(_am_cats) &
            all_active['Tier'].isin(_am_tiers)
        ].copy()

        if _am_pool.empty:
            st.warning("No active SKUs match the selected filters. Try including more categories or tiers.")
        else:
            # ── score ─────────────────────────────────────────────────────
            if _am_mode == "Revenue-weighted":
                _am_pool['_score'] = _am_pool['Daily Vel'] * _am_pool.get('Unit Price', pd.Series(0, index=_am_pool.index)).fillna(0)
            else:
                _am_pool['_score'] = _am_pool['Daily Vel']

            _am_pool = _am_pool.sort_values('_score', ascending=False).reset_index(drop=True)

            # ── selection ─────────────────────────────────────────────────
            _n = min(_am_n, len(_am_pool))

            if _am_mode == "Velocity + Coverage":
                _cat_vel   = _am_pool.groupby('Classification')['_score'].sum()
                _cat_share = (_cat_vel / _cat_vel.sum()).fillna(0)
                _cat_min   = (_cat_share * _n).round().astype(int).clip(lower=1)
                # fix rounding drift
                _drift = _n - int(_cat_min.sum())
                for _cn in (_cat_vel.sort_values(ascending=False).index if _drift > 0
                             else _cat_vel.sort_values(ascending=True).index):
                    if _drift == 0: break
                    _cat_min[_cn] += (1 if _drift > 0 else -1)
                    if _cat_min[_cn] < 1: _cat_min[_cn] = 1
                    _drift += (-1 if _drift > 0 else 1)

                _sel_skus = set()
                _sel_rows = []
                for _cn, _mn in _cat_min.items():
                    for _, _r in _am_pool[_am_pool['Classification'] == _cn].head(_mn).iterrows():
                        if _r['Supplier Sku'] not in _sel_skus:
                            _sel_skus.add(_r['Supplier Sku'])
                            _sel_rows.append(_r)
                _rem = _n - len(_sel_rows)
                if _rem > 0:
                    for _, _r in _am_pool.iterrows():
                        if _rem <= 0: break
                        if _r['Supplier Sku'] not in _sel_skus:
                            _sel_skus.add(_r['Supplier Sku'])
                            _sel_rows.append(_r)
                            _rem -= 1
                _am_menu = pd.DataFrame(_sel_rows).sort_values('_score', ascending=False).reset_index(drop=True)
            else:
                _am_menu = _am_pool.head(_n).copy().reset_index(drop=True)

            # ── status column ─────────────────────────────────────────────
            _am_menu['Status'] = _am_menu['In Stock Qty'].apply(
                lambda x: 'On Shelf' if x > 0 else 'Order Now'
            )

            # ── pre-compute gaps + costs (needed for metrics + table) ────────
            _am_gaps_early = _am_menu[_am_menu['In Stock Qty'] == 0].copy()
            _am_gaps_early['Cases'] = 1
            if 'Unit Price' in _am_gaps_early.columns and 'Pack Size' in _am_gaps_early.columns:
                _am_gaps_early['Est. Case Cost'] = (
                    _am_gaps_early['Pack Size'] * _am_gaps_early['Unit Price'].fillna(0)
                ).round(2)
            else:
                _am_gaps_early['Est. Case Cost'] = 0.0
            _am_order_cost = _am_gaps_early['Est. Case Cost'].sum()

            # ── summary metrics ───────────────────────────────────────────
            _am_on   = int((_am_menu['In Stock Qty'] > 0).sum())
            _am_off  = len(_am_menu) - _am_on
            _am_wku  = _am_menu['Daily Vel'].sum() * 7
            _am_cats_n = _am_menu['Classification'].nunique()

            st.markdown("---")
            _mc1, _mc2, _mc3, _mc4, _mc5, _mc6 = st.columns(6)
            _mc1.metric("Recommended SKUs",  f"{len(_am_menu)}")
            _mc2.metric("Already on Shelf",  f"{_am_on}")
            _mc3.metric("Need to Order",     f"{_am_off}")
            _mc4.metric("Categories",        f"{_am_cats_n}")
            _mc5.metric("Est. Weekly Units", f"{_am_wku:,.0f}")
            _mc6.metric("Est. Order Cost",   f"${_am_order_cost:,.2f}")

            st.markdown("---")

            # ── category bar chart ────────────────────────────────────────
            _cs = _am_menu.groupby('Classification').agg(
                SKUs=('Supplier Sku','count'),
                On_Shelf=('In Stock Qty', lambda x: (x > 0).sum()),
                Total_Vel=('Daily Vel','sum'),
            ).reset_index()
            _cs['Order_Now'] = _cs['SKUs'] - _cs['On_Shelf']
            # add per-category order cost
            if not _am_gaps_early.empty and 'Est. Case Cost' in _am_gaps_early.columns:
                _cat_cost = _am_gaps_early.groupby('Classification')['Est. Case Cost'].sum().rename('Order_Cost')
                _cs = _cs.merge(_cat_cost, on='Classification', how='left').fillna({'Order_Cost': 0.0})
            else:
                _cs['Order_Cost'] = 0.0
            _cs = _cs.sort_values('Total_Vel', ascending=False)

            _fig_bar = _go4.Figure()
            _fig_bar.add_bar(x=_cs['Classification'], y=_cs['On_Shelf'],
                             name='On Shelf', marker_color='#3fb950')
            _fig_bar.add_bar(x=_cs['Classification'], y=_cs['Order_Now'],
                             name='Order Now', marker_color='#f85149')
            _fig_bar.update_layout(
                barmode='stack', title='Recommended SKUs by Category',
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#c9d1d9', size=11),
                legend=dict(orientation='h', y=-0.25),
                margin=dict(t=40, b=50, l=0, r=0), height=270,
                xaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
                yaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
            )
            st.plotly_chart(_fig_bar, use_container_width=True)

            # ── category breakdown table ──────────────────────────────────
            _cs_disp = _cs[['Classification','SKUs','On_Shelf','Order_Now','Order_Cost']].rename(columns={
                'Classification': 'Category',
                'SKUs':           'Total SKUs',
                'On_Shelf':       'On Shelf',
                'Order_Now':      'Need to Order',
                'Order_Cost':     'Est. Order Cost',
            }).copy()
            st.dataframe(
                _cs_disp,
                hide_index=True,
                use_container_width=True,
                column_config={
                    'Total SKUs':      st.column_config.NumberColumn('Total SKUs',      format='%d'),
                    'On Shelf':        st.column_config.NumberColumn('On Shelf',        format='%d'),
                    'Need to Order':   st.column_config.NumberColumn('Need to Order',   format='%d'),
                    'Est. Order Cost': st.column_config.NumberColumn('Est. Order Cost', format='$%.2f'),
                }
            )
            st.caption(f"**Total: {len(_am_menu)} SKUs · ${_am_order_cost:,.2f} to order**")

            # ── velocity heatmap (category × size) ────────────────────────
            _hm = (_am_menu.groupby(['Classification','Product Size'])['Daily Vel']
                   .sum().reset_index())
            if not _hm.empty:
                _hm_piv = _hm.pivot_table(
                    index='Product Size', columns='Classification',
                    values='Daily Vel', fill_value=0
                )
                # sort sizes by the sort_size_key logic if available
                try:
                    import re as _re4
                    def _ssk4(s):
                        m = _re4.match(r'^(\d+)x(\d+\.?\d*)g?$', str(s))
                        if m: return (0, int(m.group(1)), float(m.group(2)))
                        m = _re4.match(r'^(\d+\.?\d*)g$', str(s))
                        if m: return (1, 0, float(m.group(1)))
                        m = _re4.match(r'^(\d+\.?\d*)ml$', str(s))
                        if m: return (2, 0, float(m.group(1)))
                        m = _re4.match(r'^(\d+\.?\d*)mg$', str(s))
                        if m: return (3, 0, float(m.group(1)))
                        return (99, 0, 0)
                    _hm_piv = _hm_piv.reindex(sorted(_hm_piv.index, key=_ssk4))
                except Exception:
                    pass

                _fig_hm = _go4.Figure(data=_go4.Heatmap(
                    z=_hm_piv.values,
                    x=_hm_piv.columns.tolist(),
                    y=_hm_piv.index.tolist(),
                    colorscale=[[0,'#0d1117'],[0.3,'#1f4e79'],[0.7,'#2e75b6'],[1,'#89d4f5']],
                    showscale=True,
                    hovertemplate='%{y} · %{x}<br>Total Daily Vel: %{z:.2f}<extra></extra>',
                ))
                _fig_hm.update_layout(
                    title='Velocity Heat Map — Where the Sales Are (Category × Size)',
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#c9d1d9', size=10),
                    margin=dict(t=45, b=20, l=80, r=20), height=380,
                )
                st.plotly_chart(_fig_hm, use_container_width=True)

            # ── tier distribution donut ───────────────────────────────────
            _tier_counts = _am_menu['Tier'].value_counts().reindex(['A','B','C','D']).fillna(0)
            _tier_colors = {'A':'#3fb950','B':'#89d4f5','C':'#ffc83c','D':'#f85149'}
            _fig_pie = _go4.Figure(data=_go4.Pie(
                labels=[f'Tier {t}' for t in _tier_counts.index],
                values=_tier_counts.values,
                hole=0.55,
                marker_colors=[_tier_colors.get(t,'#888') for t in _tier_counts.index],
                textinfo='label+percent',
                hovertemplate='%{label}: %{value} SKUs<extra></extra>',
            ))
            _fig_pie.update_layout(
                title='Tier Mix of Recommended Menu',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#c9d1d9', size=11),
                margin=dict(t=40, b=10, l=0, r=0),
                height=260,
                showlegend=False,
            )
            _fig_pie_col, _ = st.columns([1, 2])
            _fig_pie_col.plotly_chart(_fig_pie, use_container_width=True)

            st.markdown("---")

            # ── full recommended menu table ───────────────────────────────
            st.markdown("#### 🏆 Recommended Menu — Full List")
            _disp_c = [c for c in ['Product','Classification','Product Size','Strain','Tier',
                                    'Weekly Vel','Daily Vel','In Stock Qty','Status']
                       if c in _am_menu.columns]
            _am_disp = _am_menu[_disp_c].copy().rename(
                columns={'Classification':'Category','Product Size':'Size',
                         'In Stock Qty':'On Shelf'}
            )
            _am_disp['Daily Vel']  = _am_disp['Daily Vel'].round(3)
            _am_disp['Weekly Vel'] = _am_disp['Weekly Vel'].round(2)
            st.dataframe(
                _am_disp,
                hide_index=True,
                use_container_width=True,
                column_config={
                    'Status':     st.column_config.TextColumn('Status', width='small'),
                    'Tier':       st.column_config.TextColumn('Tier',   width='small'),
                    'Daily Vel':  st.column_config.NumberColumn('Daily Vel',  format='%.3f'),
                    'Weekly Vel': st.column_config.NumberColumn('Weekly Vel', format='%.2f'),
                    'On Shelf':   st.column_config.NumberColumn('On Shelf',   format='%d'),
                }
            )

            # ── gaps to order ─────────────────────────────────────────────
            _am_gaps = _am_gaps_early  # already computed above
            if not _am_gaps.empty:
                st.markdown("---")
                st.markdown(f"#### 📦 {len(_am_gaps)} Missing SKUs — Order to Complete Your Menu")
                st.caption(
                    "These top performers belong on your menu but aren't currently on your shelf. "
                    "Sorted by velocity — the top of this list is your biggest missed opportunity."
                )
                _gap_c = [c for c in ['Product','Classification','Product Size','Strain','Tier',
                                       'Weekly Vel','Daily Vel','Cases','Unit Price','Est. Case Cost']
                          if c in _am_gaps.columns]
                _gap_disp = _am_gaps[_gap_c].copy().rename(
                    columns={'Classification':'Category','Product Size':'Size'}
                )
                _gap_disp['Daily Vel']  = _gap_disp['Daily Vel'].round(3)
                _gap_disp['Weekly Vel'] = _gap_disp['Weekly Vel'].round(2)
                st.dataframe(
                    _gap_disp,
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        'Tier':           st.column_config.TextColumn('Tier',   width='small'),
                        'Cases':          st.column_config.NumberColumn('Cases', format='%d', width='small'),
                        'Daily Vel':      st.column_config.NumberColumn('Daily Vel',      format='%.3f'),
                        'Weekly Vel':     st.column_config.NumberColumn('Weekly Vel',     format='%.2f'),
                        'Unit Price':     st.column_config.NumberColumn('Unit Price',     format='$%.2f'),
                        'Est. Case Cost': st.column_config.NumberColumn('Est. Case Cost', format='$%.2f'),
                    }
                )
                st.caption(f"{len(_am_gaps)} SKUs · {int(_am_gaps['Cases'].sum())} cases · **Est. ${_am_order_cost:,.2f}**")

            # ── download ──────────────────────────────────────────────────
            st.markdown("---")
            _ai_buf = io.BytesIO()
            with pd.ExcelWriter(_ai_buf, engine='openpyxl') as _aiw:
                # Sheet 1 — full recommended menu
                _exp_cols = [c for c in ['Product','Classification','Product Size','Strain','Tier',
                                          'Weekly Vel','Daily Vel','In Stock Qty','Unit Price','Supplier Sku','Status']
                             if c in _am_menu.columns]
                _menu_sheet = _am_menu[_exp_cols].rename(
                    columns={'Classification':'Category','Product Size':'Size','In Stock Qty':'On Shelf'}
                ).copy()
                _menu_sheet['Daily Vel']  = _menu_sheet['Daily Vel'].round(3)
                _menu_sheet['Weekly Vel'] = _menu_sheet['Weekly Vel'].round(2)
                _menu_sheet.to_excel(_aiw, sheet_name='AI Recommended Menu', index=False, startrow=5)
                # Write filter metadata above the table
                _ws_menu = _aiw.sheets['AI Recommended Menu']
                _ws_menu.cell(row=1, column=1, value='AI Menu Optimizer — Recommended Menu').font = Font(bold=True, size=12)
                _ws_menu.cell(row=2, column=1, value=f'Generated: {_today4.strftime("%B %d, %Y")}')
                _ws_menu.cell(row=3, column=1, value=f'Mode: {_am_mode}  |  Target SKUs: {_am_n}  |  Tiers: {", ".join(_am_tiers)}')
                _ws_menu.cell(row=4, column=1, value=f'Categories: {", ".join(_am_cats)}')

                # Sheet 2 — gaps to order (always included)
                _gap_exp_cols = [c for c in ['Product','Classification','Product Size','Strain','Tier',
                                              'Weekly Vel','Daily Vel','Cases','Unit Price','Est. Case Cost','Supplier Sku']
                                 if c in _am_gaps.columns]
                if not _am_gaps.empty and _gap_exp_cols:
                    _gap_sheet = _am_gaps[_gap_exp_cols].rename(
                        columns={'Classification':'Category','Product Size':'Size'}
                    ).copy()
                    _gap_sheet['Daily Vel']  = _gap_sheet['Daily Vel'].round(3)
                    _gap_sheet['Weekly Vel'] = _gap_sheet['Weekly Vel'].round(2)
                    _gap_sheet.to_excel(_aiw, sheet_name='Order to Complete Menu', index=False, startrow=4)
                    _ws_gap = _aiw.sheets['Order to Complete Menu']
                    _ws_gap.cell(row=1, column=1, value='Order to Complete AI Menu').font = Font(bold=True, size=12)
                    _ws_gap.cell(row=2, column=1, value=f'Generated: {_today4.strftime("%B %d, %Y")}  |  Mode: {_am_mode}  |  Tiers: {", ".join(_am_tiers)}')
                    _ws_gap.cell(row=3, column=1, value=f'Categories: {", ".join(_am_cats)}')
                else:
                    pd.DataFrame([{'Note': 'All recommended SKUs are already on your shelf — no gaps to fill.'}]).to_excel(
                        _aiw, sheet_name='Order to Complete Menu', index=False)

                # Sheet 3 — category breakdown
                _cs.rename(columns={'On_Shelf':'On Shelf','Order_Now':'Order Now',
                                    'Total_Vel':'Total Daily Vel','Order_Cost':'Est. Order Cost'}).to_excel(
                    _aiw, sheet_name='Category Breakdown', index=False)

            _ai_buf.seek(0)

            # Separate OCS upload file — must be its own file with MasterCatalogue
            # as the FIRST (and only) sheet so the OCS portal can read it
            _ocs_buf = io.BytesIO()
            if not _am_gaps.empty and 'Supplier Sku' in _am_gaps.columns:
                _ocs_df = _am_gaps[['Supplier Sku','Cases']].rename(
                    columns={'Supplier Sku':'SKU','Cases':'Quantity'}
                ).copy()
                with pd.ExcelWriter(_ocs_buf, engine='openpyxl') as _ocs_w:
                    _ocs_df.to_excel(_ocs_w, sheet_name='MasterCatalogue', index=False)
            _ocs_buf.seek(0)

            _dl_c1, _dl_c2 = st.columns(2)
            with _dl_c1:
                st.download_button(
                    "📥 Download AI Menu Report (.xlsx)",
                    data=_ai_buf,
                    file_name=f'AIMenu_{_today4.strftime("%Y%m%d")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                    help="Full report: recommended menu, gaps to order, category breakdown.",
                )
            with _dl_c2:
                st.download_button(
                    "📤 Upload to OCS (.xlsx)",
                    data=_ocs_buf,
                    file_name=f'OCS_AIMenu_{_today4.strftime("%Y%m%d")}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                    help="Import this file directly into the OCS portal — SKU + Quantity only.",
                    disabled=_am_gaps.empty,
                )
            st.caption(
                f"**{len(_am_menu)} SKUs** recommended &nbsp;·&nbsp; "
                f"**{len(_am_gaps)} to order** &nbsp;·&nbsp; "
                f"**{_am_on} on shelf** &nbsp;·&nbsp; "
                f"**Est. ${_am_order_cost:,.2f}** &nbsp;·&nbsp; "
                f"Mode: {_am_mode}",
                unsafe_allow_html=True,
            )
